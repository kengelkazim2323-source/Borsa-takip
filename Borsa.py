import streamlit as st
import yfinance as yf
import pandas as pd
import json
import os
import pytz
from datetime import datetime, timedelta
import plotly.express as px
import plotly.graph_objects as go
import numpy as np
import time as _time
from streamlit_autorefresh import st_autorefresh
import urllib.request
import re
from concurrent.futures import ThreadPoolExecutor
import logging

logging.basicConfig(level=logging.WARNING, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==========================================
# 0. VERİ YÖNETİMİ
# ==========================================
# Dosyaları daima bu script'in bulunduğu dizine kaydet.
# Böylece kod güncellendiğinde veya farklı dizinden çalıştırıldığında
# veriler kaybolmaz.
_VERI_DIZIN = os.path.dirname(os.path.abspath(__file__))

def _veri_yolu(dosya_adi):
    """Dosya adını, script'in dizinine göre mutlak yola çevirir."""
    return os.path.join(_VERI_DIZIN, dosya_adi)

PORTFOY_DOSYASI    = _veri_yolu("portfoy_kayitlari.json")
IPO_DOSYASI        = _veri_yolu("halka_arz_kayitlari.json")
ALARM_DOSYASI      = _veri_yolu("alarm_kayitlari.json")
PERFORMANS_DOSYASI = _veri_yolu("portfoy_performans.json")
NOTLAR_DOSYASI     = _veri_yolu("hisse_notlari.json")

def load_json(dosya_adi):
    """
    JSON dosyasını yükler.
    Dosya bozuksa .bak yedeğini dener.
    Veri yoksa boş liste döner — asla exception fırlatmaz.
    """
    for deneme in [dosya_adi, dosya_adi + ".bak"]:
        if not os.path.exists(deneme):
            continue
        try:
            with open(deneme, "r", encoding="utf-8") as f:
                data = json.load(f)
                if not isinstance(data, list):
                    continue
                # Performans dosyası 'tarih' key'i kullanır
                if data and 'tarih' in data[0] and 'Hisse' not in data[0]:
                    return sorted(data, key=lambda x: x.get('tarih', ''))
                return sorted(data, key=lambda x: x.get('Hisse', ''))
        except Exception as e:
            logger.error(f"JSON yükleme hatası ({deneme}): {e}")
    return []

def save_json(dosya_adi, data):
    """
    JSON dosyasını kaydeder.
    Kaydetmeden önce mevcut dosyayı .bak olarak yedekler.
    Atomic write: önce .tmp'ye yazar, sonra rename eder — yarım kayıt olmaz.
    """
    try:
        # Varsa mevcut dosyayı yedekle
        if os.path.exists(dosya_adi):
            try:
                import shutil
                shutil.copy2(dosya_adi, dosya_adi + ".bak")
            except Exception as e:
                logger.warning(f"Yedekleme başarısız ({dosya_adi}): {e}")

        # Geçici dosyaya yaz → atomic rename
        tmp = dosya_adi + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        os.replace(tmp, dosya_adi)   # atomic: yarım kayıt imkansız

    except Exception as e:
        logger.error(f"JSON kaydetme hatası ({dosya_adi}): {e}")
        # Geçici dosya kaldıysa temizle
        try:
            if os.path.exists(dosya_adi + ".tmp"):
                os.remove(dosya_adi + ".tmp")
        except Exception:
            pass

if 'portfoy' not in st.session_state:
    st.session_state.portfoy = load_json(PORTFOY_DOSYASI)
if 'ipo_liste' not in st.session_state:
    st.session_state.ipo_liste = load_json(IPO_DOSYASI)
if 'alarmlar' not in st.session_state:
    st.session_state.alarmlar = load_json(ALARM_DOSYASI)
if 'performans' not in st.session_state:
    st.session_state.performans = load_json(PERFORMANS_DOSYASI)
if 'notlar' not in st.session_state:
    raw_notlar = load_json(NOTLAR_DOSYASI)
    st.session_state.notlar = {n['Hisse']: n for n in raw_notlar} if raw_notlar else {}
if 'dark_mode' not in st.session_state:
    st.session_state.dark_mode = True

@st.cache_data(ttl=3600)
def fetch_temel_veri(symbol):
    """yfinance'tan F/K, PD/DD, Piyasa Değeri ve Borç/Özsermaye oranını çeker."""
    code = symbol.replace(".IS", "").upper()
    try:
        tk   = yf.Ticker(symbol)
        info = tk.info
        if not info:
            return None
        def safe(key, fallback=None):
            v = info.get(key)
            return v if v is not None else fallback

        pd_dd  = safe('priceToBook')
        fk     = safe('trailingPE') or safe('forwardPE')
        borc   = safe('totalDebt', 0)
        ozserm = safe('totalStockholderEquity') or safe('bookValue', 0)
        borc_ozserm = round(borc / ozserm, 2) if ozserm and ozserm > 0 else None
        piyasa_degeri = safe('marketCap')
        temttu_verimi = safe('dividendYield')

        return {
            'FK':           round(fk, 2)            if fk          else None,
            'PD_DD':        round(pd_dd, 2)          if pd_dd       else None,
            'Borc_Ozserm':  borc_ozserm,
            'Piyasa_Degeri': piyasa_degeri,
            'Temettu_Verimi': round(temttu_verimi * 100, 2) if temttu_verimi else None,
        }
    except Exception as e:
        logger.warning(f"Temel veri hatası ({code}): {e}")
        return None


@st.cache_data(ttl=1800)
def fetch_haberler(hisse_listesi):
    """Portföydeki hisseler için yfinance news API'sinden son haberleri çeker."""
    haberler = []
    for symbol in hisse_listesi[:8]:   # max 8 hisse, yük azaltmak için
        code = symbol.replace(".IS", "")
        try:
            tk   = yf.Ticker(symbol)
            news = tk.news or []
            for item in news[:3]:      # hisse başına en fazla 3 haber
                haberler.append({
                    'hisse':   code,
                    'baslik':  item.get('title', ''),
                    'url':     item.get('link', ''),
                    'kaynak':  item.get('publisher', ''),
                    'zaman':   item.get('providerPublishTime', 0),
                })
        except Exception as e:
            logger.warning(f"Haber çekme hatası ({code}): {e}")
    # Zamana göre sırala (en yeni önce)
    haberler.sort(key=lambda x: x['zaman'], reverse=True)
    return haberler


@st.cache_data(ttl=300)
def fetch_bist100_karsilastirma(portfoy_hisseleri):
    """BIST100 ve portföy günlük getiri verilerini karşılaştırma için çeker."""
    try:
        bist = fetch_stock_data("XU100.IS")
        if not bist:
            return None, None
        bist_close = bist['hist']['Close']

        # Portföy ağırlıklı ortalama getirisi
        portfoy_close = {}
        for h in portfoy_hisseleri:
            d = fetch_stock_data(h)
            if d and not d['hist'].empty:
                portfoy_close[h] = d['hist']['Close']

        if not portfoy_close:
            return bist_close, None

        # Ortak tarih aralığına normalize et
        min_len = min(len(bist_close), min(len(v) for v in portfoy_close.values()))
        bist_norm = (bist_close.iloc[-min_len:] / bist_close.iloc[-min_len] * 100)

        portfoy_df = pd.DataFrame({h: v.iloc[-min_len:].values for h, v in portfoy_close.items()})
        portfoy_ort = portfoy_df.mean(axis=1)
        portfoy_norm = (portfoy_ort / portfoy_ort.iloc[0] * 100)
        portfoy_norm.index = bist_close.iloc[-min_len:].index

        return bist_norm, portfoy_norm
    except Exception as e:
        logger.warning(f"Karşılaştırma hatası: {e}")
        return None, None


@st.cache_data(ttl=3600)
def fetch_temettu_isyatirim(symbol):
    """isyatirim.com.tr API'sinden net temettü verisini çeker (birincil kaynak)."""
    code = symbol.replace(".IS", "").upper()
    try:
        ts       = int(_time.time() * 1000)
        bugun    = datetime.now()
        bitis    = bugun.strftime("%d-%m-%Y")
        baslangic = (bugun - timedelta(days=730)).strftime("%d-%m-%Y")
        url = (
            f"https://www.isyatirim.com.tr/_layouts/15/IsYatirim.Website/Common/Data.aspx/"
            f"HisseTemettuTablosu?hisse={code}&baslangic={baslangic}&bitis={bitis}&_={ts}"
        )
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
            'Accept': 'application/json, text/javascript, */*; q=0.01',
            'X-Requested-With': 'XMLHttpRequest',
            'Referer': f'https://www.isyatirim.com.tr/analiz-ve-raporlar/analiz/hisse/{code}',
        })
        response = urllib.request.urlopen(req, timeout=8)
        data = json.loads(response.read().decode('utf-8'))

        kayitlar = data.get('value') or data.get('Value') or []
        if not kayitlar:
            return None

        son_1_yil  = datetime.now() - timedelta(days=365)
        son_2_yil  = datetime.now() - timedelta(days=730)
        parsed = []
        for item in kayitlar:
            # Net temettü tutarı — farklı field isimlerine karşı güvenli
            net = (item.get('NET_TEMETTU_TUTARI') or item.get('NetTemettuTutari')
                   or item.get('netTemettuTutari') or item.get('NetTemettu') or 0)
            brut = (item.get('BRUT_TEMETTU_TUTARI') or item.get('BrutTemettuTutari')
                    or item.get('brutTemettuTutari') or item.get('BrutTemettu') or 0)
            tarih_str = (item.get('TEMETTU_ODEME_TARIHI') or item.get('OdemeTarihi')
                         or item.get('Tarih') or item.get('tarih') or '')
            try:
                net_val  = float(str(net).replace(',', '.').replace(' ', ''))
                brut_val = float(str(brut).replace(',', '.').replace(' ', ''))
                if not tarih_str:
                    continue
                # Tarih formatları: dd.mm.yyyy veya yyyy-mm-dd
                for fmt in ('%d.%m.%Y', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d'):
                    try:
                        dt = datetime.strptime(tarih_str[:10], fmt)
                        break
                    except ValueError:
                        dt = None
                if dt and net_val > 0:
                    parsed.append({'dt': dt, 'net': net_val, 'brut': brut_val,
                                   'tarih_str': dt.strftime('%d.%m.%Y')})
            except Exception:
                continue

        if not parsed:
            return None

        parsed.sort(key=lambda x: x['dt'], reverse=True)
        son_1_yil_list = [p for p in parsed if p['dt'] >= son_1_yil]
        son_2_yil_list = [p for p in parsed if p['dt'] >= son_2_yil]

        if son_1_yil_list:
            net_toplam = sum(p['net'] for p in son_1_yil_list)
            n = len(son_1_yil_list)
            son_tarih = son_1_yil_list[0]['tarih_str'] + (f" ({n}x)" if n > 1 else "")
            return {'net': round(net_toplam, 6), 'tarih': son_tarih, 'kaynak': 'isyatirim'}
        elif son_2_yil_list:
            net_val = son_2_yil_list[0]['net']
            son_tarih = son_2_yil_list[0]['tarih_str'] + " *"
            return {'net': round(net_val, 6), 'tarih': son_tarih, 'kaynak': 'isyatirim'}

        return None
    except Exception as e:
        logger.warning(f"isyatirim temettü hatası ({code}): {e}")
        return None


def piyasa_acik_mi():
    """BIST'in şu an açık olup olmadığını İstanbul saatine göre döner."""
    try:
        tz      = pytz.timezone('Europe/Istanbul')
        su_an   = datetime.now(tz)
        gun     = su_an.weekday()      # 0=Pzt … 6=Paz
        saat    = su_an.hour
        dakika  = su_an.minute
        if gun >= 5:
            return False, "Hafta Sonu"
        toplam_dakika = saat * 60 + dakika
        acilis  = 10 * 60          # 10:00
        kapanis = 18 * 60          # 18:00
        if acilis <= toplam_dakika <= kapanis:
            return True, f"Açık · Kapanışa {18*60 - toplam_dakika} dk"
        elif toplam_dakika < acilis:
            return False, f"Kapalı · {10*60 - toplam_dakika} dk'ya açılıyor"
        else:
            return False, "Kapandı"
    except Exception:
        return None, "?"


def kaydet_performans_snapshot(full_data_list):
    """Her yüklemede o günkü portföy değerini performans geçmişine kaydeder."""
    if not full_data_list:
        return
    try:
        bugun       = datetime.now(pytz.timezone('Europe/Istanbul')).strftime('%Y-%m-%d')
        toplam_deger = round(sum(x['Değer'] for x in full_data_list), 2)
        toplam_kz   = round(sum(x['K/Z']   for x in full_data_list), 2)

        kayitlar = load_json(PERFORMANS_DOSYASI)
        # list formatına dönüştür (load_json sıralı döndürür, biz dict listesi tutuyoruz)
        if not isinstance(kayitlar, list):
            kayitlar = []

        guncellendi = False
        for k in kayitlar:
            if k.get('tarih') == bugun:
                k['deger'] = toplam_deger
                k['kz']    = toplam_kz
                guncellendi = True
                break
        if not guncellendi:
            kayitlar.append({'tarih': bugun, 'deger': toplam_deger, 'kz': toplam_kz})

        kayitlar = sorted(kayitlar, key=lambda x: x.get('tarih', ''))[-365:]
        save_json(PERFORMANS_DOSYASI, kayitlar)
        st.session_state.performans = kayitlar
    except Exception as e:
        logger.error(f"Performans kayıt hatası: {e}")


# ==========================================
# VERİ ÇEKME FONKSİYONLARI
# ==========================================
@st.cache_data(ttl=300)
def fetch_stock_data(symbol):
    try:
        tk = yf.Ticker(symbol)
        hist = tk.history(period="60d")   # MACD için 60 gün gerekli
        if hist.empty:
            logger.warning(f"Veri boş: {symbol}")
            return None

        divs = tk.dividends
        yillik_net_temettu = 0.0
        son_tarih = "-"

        if not divs.empty:
            try:
                # Timezone sorununu düzelt: tz-aware → tz-naive
                if divs.index.tz is not None:
                    divs.index = divs.index.tz_convert('UTC').tz_localize(None)
                else:
                    divs.index = divs.index.tz_localize(None)

                bugun          = datetime.now()
                son_1_yil      = bugun - timedelta(days=365)
                son_2_yil      = bugun - timedelta(days=730)
                son_1_yil_divs = divs[divs.index >= son_1_yil]
                son_2_yil_divs = divs[divs.index >= son_2_yil]

                if not son_1_yil_divs.empty:
                    # Son 1 yılda dağıtım var → tümünü topla (brüt)
                    yillik_brut_temettu = float(son_1_yil_divs.sum())
                    yillik_net_temettu  = round(yillik_brut_temettu * 0.90, 6)
                    son_tarih = divs.index[-1].strftime('%d.%m.%Y')
                    # Birden fazla dağıtım varsa "(Nx)" göster
                    n = len(son_1_yil_divs)
                    if n > 1:
                        son_tarih += f" ({n}x)"
                elif not son_2_yil_divs.empty:
                    # 1-2 yıl arası dağıtım var → en güncel tane
                    yillik_brut_temettu = float(divs.iloc[-1])
                    yillik_net_temettu  = round(yillik_brut_temettu * 0.90, 6)
                    son_tarih = divs.index[-1].strftime('%d.%m.%Y') + " *"
                else:
                    # Hiç son verisi yok
                    yillik_net_temettu = 0.0
                    son_tarih = "-"
            except Exception as e:
                logger.warning(f"Temettü işleme hatası ({symbol}): {e}")
                yillik_net_temettu = 0.0
                son_tarih = "-"

        # isyatirim.com.tr birincil kaynak olarak dene (daha doğru net temettü)
        isy = fetch_temettu_isyatirim(symbol)
        if isy:
            yillik_net_temettu = isy['net']
            son_tarih          = isy['tarih']

        return {"hist": hist, "temettu": yillik_net_temettu, "tarih": son_tarih}

    except Exception as e:
        logger.error(f"Hisse veri çekme hatası ({symbol}): {e}")
        return None

@st.cache_data(ttl=300)
def fetch_tefas_price(symbol):
    """TEFAS'tan fon fiyatı çeker. Çoklu regex ve hata durumu döner."""
    code = symbol.replace(".IS", "").upper()
    url = f"https://www.tefas.gov.tr/FonAnaliz.aspx?FonKod={code}"
    try:
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        html = urllib.request.urlopen(req, timeout=8).read().decode('utf-8')

        # Sırayla farklı regex dene
        patterns = [
            r'Son Fiyat \(TL\).*?<span[^>]*>([\d.,]+)</span>',
            r'top-list-right">([\d.,]+)</span>',
            r'"top-list-right">\s*([\d.,]+)',
        ]
        for pattern in patterns:
            m = re.search(pattern, html, re.DOTALL)
            if m:
                price_str = m.group(1).strip().replace('.', '').replace(',', '.')
                try:
                    price = float(price_str)
                    if price > 0:
                        return {"fiyat": price, "durum": "ok"}
                except ValueError:
                    continue

        logger.warning(f"TEFAS fiyat parse edilemedi: {code}")
        return {"fiyat": None, "durum": "parse_hatası"}

    except urllib.error.URLError as e:
        logger.error(f"TEFAS bağlantı hatası ({code}): {e}")
        return {"fiyat": None, "durum": "bağlantı_hatası"}
    except Exception as e:
        logger.error(f"TEFAS genel hata ({code}): {e}")
        return {"fiyat": None, "durum": "genel_hata"}

def tr_format(val):
    try:
        if val is None or pd.isna(val): return "0,00"
        return f"{val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except: return "0,00"

def tr_format4(val):
    """Maliyet gibi hassas değerler için 4 ondalık basamak."""
    try:
        if val is None or pd.isna(val): return "0,0000"
        return f"{val:,.4f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except: return "0,0000"

def hex_rgba(hex_color, alpha=0.08):
    """6-digit hex rengi plotly uyumlu rgba() string'e çevirir."""
    try:
        h = hex_color.lstrip('#')
        r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        return f"rgba({r},{g},{b},{alpha})"
    except Exception:
        return f"rgba(128,128,128,{alpha})"

# ==========================================
# GELİŞTİRİLMİŞ SİNYAL MOTORU
# RSI + MA20 + MACD + Bollinger Bands
# ==========================================
def get_signal(hist_data):
    try:
        if len(hist_data) < 26:
            return "VERİ YETERSİZ", 0.0, 0.0, 50.0

        close = hist_data['Close']

        # --- RSI (14) ---
        delta = close.diff()
        gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
        loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
        rs = gain / loss
        rsi = float(100 - (100 / (1 + rs)).iloc[-1])

        # --- MA20 ---
        ma20 = float(close.rolling(window=20).mean().iloc[-1])
        last_price = float(close.iloc[-1])

        # --- MACD (12, 26, 9) ---
        ema12 = close.ewm(span=12, adjust=False).mean()
        ema26 = close.ewm(span=26, adjust=False).mean()
        macd_line = ema12 - ema26
        signal_line = macd_line.ewm(span=9, adjust=False).mean()
        macd_val = float(macd_line.iloc[-1])
        signal_val = float(signal_line.iloc[-1])
        macd_histogram = round(macd_val - signal_val, 4)

        # --- Bollinger Bands (20, 2σ) ---
        bb_ma = close.rolling(window=20).mean()
        bb_std = close.rolling(window=20).std()
        bb_upper = float((bb_ma + 2 * bb_std).iloc[-1])
        bb_lower = float((bb_ma - 2 * bb_std).iloc[-1])
        bb_pct = round(
            (last_price - bb_lower) / (bb_upper - bb_lower) * 100, 1
        ) if bb_upper != bb_lower else 50.0

        # --- Birleşik Skor ---
        skor = 0

        # RSI katkısı
        if rsi < 30:   skor += 3
        elif rsi < 40: skor += 2
        elif rsi < 50: skor += 1
        elif rsi > 75: skor -= 3
        elif rsi > 65: skor -= 2
        elif rsi > 55: skor -= 1

        # MA20 katkısı
        if last_price > ma20:  skor += 1
        else:                   skor -= 1

        # MACD katkısı (histogram yönü)
        if macd_val > signal_val:  skor += 2
        else:                       skor -= 2

        # Bollinger katkısı
        if last_price < bb_lower:   skor += 2   # Aşırı satım
        elif last_price > bb_upper: skor -= 2   # Aşırı alım

        # Sinyal kararı
        if   skor >= 6:  sinyal = "🟢 GÜÇLÜ AL"
        elif skor >= 3:  sinyal = "🟢 AL"
        elif skor <= -6: sinyal = "🔴 GÜÇLÜ SAT"
        elif skor <= -3: sinyal = "🔴 SAT"
        else:            sinyal = "🟡 TUT"

        return sinyal, round(rsi, 1), macd_histogram, bb_pct

    except Exception as e:
        logger.warning(f"Sinyal hesaplama hatası: {e}")
        return "---", 0.0, 0.0, 50.0

# ==========================================
# PARALEL VERİ HAZIRLAMA
# ==========================================
def fetch_single_item(args):
    """Tek bir portföy kalemini işler (ThreadPoolExecutor için)."""
    i, item = args
    piyasa_durumu = item.get("Piyasa", "Türk Borsası")
    d = fetch_stock_data(item['Hisse'])

    if piyasa_durumu == "Yatırım Fonu":
        tefas_result = fetch_tefas_price(item['Hisse'])
        fiyat = tefas_result.get("fiyat") if tefas_result else None
        tefas_durum = tefas_result.get("durum", "hata") if tefas_result else "hata"

        if fiyat:
            c = fiyat
            if d:
                sinyal_result = get_signal(d['hist'])
                sinyal, rsi_val, macd_h, bb_pct = sinyal_result
                pc = float(d['hist']['Close'].iloc[-2])
            else:
                sinyal = "VERİ YOK"; rsi_val = 0.0; macd_h = 0.0; bb_pct = 50.0; pc = c
        else:
            c = item['Maliyet']; pc = c
            sinyal = f"⚠️ {tefas_durum.replace('_', ' ').upper()}"
            rsi_val = 0.0; macd_h = 0.0; bb_pct = 50.0
        temettu = 0.0; tarih = "-"

    else:
        if d:
            c = float(d['hist']['Close'].iloc[-1])
            pc = float(d['hist']['Close'].iloc[-2])
            sinyal, rsi_val, macd_h, bb_pct = get_signal(d['hist'])
            temettu = d['temettu']; tarih = d['tarih']
        else:
            c = item['Maliyet']; pc = c
            sinyal = "⚠️ VERİ YOK"
            rsi_val = 0.0; macd_h = 0.0; bb_pct = 50.0
            temettu = 0.0; tarih = "-"

    adet_int = int(item['Adet'])

    # Son 7 günlük kapanış verisi (sparkline için)
    spark_prices = []
    if d and not d['hist'].empty:
        try:
            spark_prices = [round(float(v), 4) for v in d['hist']['Close'].iloc[-7:].tolist()]
        except Exception:
            spark_prices = []

    return {
        "id": i, "Piyasa": piyasa_durumu, "Hisse": item['Hisse'],
        "Sinyal": sinyal, "RSI": rsi_val, "MACD_H": macd_h, "BB_PCT": bb_pct,
        "Adet": adet_int, "Maliyet": item['Maliyet'],
        "Güncel": c,
        "K/Z": round((c - item['Maliyet']) * adet_int, 2),
        "Değer": round(c * adet_int, 2),
        "Temettu": temettu,
        "NetTemettu": round(temettu * adet_int, 2),
        "DailyDiff": round((c - pc) * adet_int, 2),
        "Tarih": tarih,
        "Sparkline": spark_prices,
    }

# ==========================================
# 1. TEMA VE CSS
# ==========================================
st.set_page_config(page_title="Borsa Takip", page_icon="📈", layout="wide")

# Yenileme süresi session_state'ten okunur (sidebar'dan ayarlanabilir)
if 'yenileme_suresi' not in st.session_state:
    st.session_state.yenileme_suresi = 60   # saniye

st_autorefresh(interval=st.session_state.yenileme_suresi * 1000, key="datarefresh")

tema_isimleri = [
    "Siyah-Beyaz (Klasik)", "Siyah-Beyaz (Koyu)", "Galaksi (VIP)", "Siber Punk", "Matrix", "Altın Vuruş",
    "Zümrüt Yeşili", "Lav Akışı", "Okyanus Derinliği", "Kuzey Işıkları", "Buzul (Dark)", "Mor Ötesi",
    "Bakır Buharı", "Gece Yarısı", "Safir Gece", "Çöl Fırtınası", "Kızıl Elmas", "Premium Koyu",
    "Retro Kehribar", "Derin Orman", "Antrasit VIP", "Neon Gecesi", "Gümüş", "Titanyum", "Platin",
    "Yakut", "Ametist", "Turkuaz", "Karanlık Madde", "Süpernova", "Karadelik", "Yıldız Tozu",
    "Kozmik Mavi", "Güneş Patlaması", "Zehirli Yeşil", "Çikolata Rüyası", "Vanilya Gökyüzü",
    "Kızıl Gezegen", "Buz Devi", "Volkanik Kül", "Şafak Vakti", "Alacakaranlık", "Gece Kuşu",
    "Şehir Işıkları", "Siber Gümüş", "Kripto Yeşili", "Hacker Terminali", "Galaktik Mor",
    "Kuantum Foton", "Biyolüminesans"
]

FONT_SECENEKLERI = {
    "Inter (Varsayılan)":    ("Inter", "https://fonts.googleapis.com/css2?family=Inter:wght@400;600&display=swap"),
    "Roboto":                ("Roboto", "https://fonts.googleapis.com/css2?family=Roboto:wght@400;700&display=swap"),
    "Poppins":               ("Poppins", "https://fonts.googleapis.com/css2?family=Poppins:wght@400;600&display=swap"),
    "Nunito":                ("Nunito", "https://fonts.googleapis.com/css2?family=Nunito:wght@400;700&display=swap"),
    "Raleway":               ("Raleway", "https://fonts.googleapis.com/css2?family=Raleway:wght@400;600&display=swap"),
    "Source Code Pro":       ("Source Code Pro", "https://fonts.googleapis.com/css2?family=Source+Code+Pro:wght@400;600&display=swap"),
    "Exo 2":                 ("Exo 2", "https://fonts.googleapis.com/css2?family=Exo+2:wght@400;600&display=swap"),
    "Orbitron (Sci-Fi)":     ("Orbitron", "https://fonts.googleapis.com/css2?family=Orbitron:wght@400;700&display=swap"),
    "Share Tech Mono":       ("Share Tech Mono", "https://fonts.googleapis.com/css2?family=Share+Tech+Mono&display=swap"),
    "Syne":                  ("Syne", "https://fonts.googleapis.com/css2?family=Syne:wght@400;700&display=swap"),
}

with st.sidebar:
    st.header("🎨 Tema Galerisi")
    # session_state ile seçim kalıcı kalır (tema değişince rerun da etkilemez)
    if 'tema_secim' not in st.session_state:
        st.session_state.tema_secim = "Galaksi (VIP)"
    if 'font_secim' not in st.session_state:
        st.session_state.font_secim = "Inter (Varsayılan)"

    tema = st.selectbox(
        "Görünüm Seç", tema_isimleri,
        index=tema_isimleri.index(st.session_state.tema_secim)
              if st.session_state.tema_secim in tema_isimleri else 2,
        key="tema_widget"
    )
    st.session_state.tema_secim = tema

    secili_font_adi = st.selectbox(
        "🔤 Font Seç", list(FONT_SECENEKLERI.keys()),
        index=list(FONT_SECENEKLERI.keys()).index(st.session_state.font_secim)
              if st.session_state.font_secim in FONT_SECENEKLERI else 0,
        key="font_widget"
    )
    st.session_state.font_secim = secili_font_adi
    secili_font, secili_font_url = FONT_SECENEKLERI[secili_font_adi]

tema_renkleri = {
    "Siyah-Beyaz (Klasik)": {"bg": "#FFFFFF", "text": "#000000", "box": "#F5F5F5", "accent": "#000000"},
    "Siyah-Beyaz (Koyu)":   {"bg": "#000000", "text": "#FFFFFF", "box": "#1A1A1A", "accent": "#FFFFFF"},
    "Galaksi (VIP)":        {"bg": "#0B0E14", "text": "#E0E0E0", "box": "#161B22", "accent": "#00D4FF"},
    "Siber Punk":           {"bg": "#0D0221", "text": "#FFFFFF", "box": "#190033", "accent": "#FF00FF"},
    "Matrix":               {"bg": "#000000", "text": "#00FF41", "box": "#0D0208", "accent": "#00FF41"},
    "Altın Vuruş":          {"bg": "#0F0F0F", "text": "#F5F5F5", "box": "#1A1A1A", "accent": "#D4AF37"},
    "Zümrüt Yeşili":        {"bg": "#06120B", "text": "#E8F5E9", "box": "#0D2114", "accent": "#00E676"},
    "Lav Akışı":            {"bg": "#1A0F0F", "text": "#F8F9FA", "box": "#2D1B1B", "accent": "#FF4D4D"},
    "Okyanus Derinliği":    {"bg": "#001B2E", "text": "#ADB5BD", "box": "#003554", "accent": "#24D1FF"},
    "Kuzey Işıkları":       {"bg": "#0B101B", "text": "#E9ECEF", "box": "#1B263B", "accent": "#A5FFD6"},
    "Buzul (Dark)":         {"bg": "#0D1117", "text": "#C9D1D9", "box": "#161B22", "accent": "#58A6FF"},
    "Mor Ötesi":            {"bg": "#120D1D", "text": "#E0D7FF", "box": "#1E1631", "accent": "#9D4EDD"},
    "Bakır Buharı":         {"bg": "#1B1510", "text": "#D4A373", "box": "#2C211A", "accent": "#E76F51"},
    "Gece Yarısı":          {"bg": "#050505", "text": "#FFFFFF", "box": "#121212", "accent": "#F72585"},
    "Safir Gece":           {"bg": "#03045E", "text": "#CAF0F8", "box": "#023E8A", "accent": "#00B4D8"},
    "Çöl Fırtınası":        {"bg": "#1C1917", "text": "#F5F5F4", "box": "#292524", "accent": "#EAB308"},
    "Kızıl Elmas":          {"bg": "#0F0202", "text": "#FFFFFF", "box": "#1F0505", "accent": "#D00000"},
    "Premium Koyu":         {"bg": "#121212", "text": "#ffffff", "box": "#4c4c4c", "accent": "#BB86FC"},
    "Retro Kehribar":       {"bg": "#0A0A0A", "text": "#FFB300", "box": "#1A1A1A", "accent": "#FF8F00"},
    "Derin Orman":          {"bg": "#081C15", "text": "#D8F3DC", "box": "#1B4332", "accent": "#95D5B2"},
    "Antrasit VIP":         {"bg": "#1B1B1B", "text": "#D1D1D1", "box": "#2D2D2D", "accent": "#E0E0E0"},
    "Neon Gecesi":          {"bg": "#000814", "text": "#FFFFFF", "box": "#001D3D", "accent": "#FFC300"},
    "Gümüş":                {"bg": "#111111", "text": "#E0E0E0", "box": "#222222", "accent": "#C0C0C0"},
    "Titanyum":             {"bg": "#1C1F22", "text": "#E8E9EA", "box": "#2B2F33", "accent": "#878681"},
    "Platin":               {"bg": "#151515", "text": "#FDFDFD", "box": "#252525", "accent": "#E5E4E2"},
    "Yakut":                {"bg": "#1A0505", "text": "#FDE0E0", "box": "#2B0A0A", "accent": "#E0115F"},
    "Ametist":              {"bg": "#140A1A", "text": "#EAD5F7", "box": "#251330", "accent": "#9966CC"},
    "Turkuaz":              {"bg": "#061A1C", "text": "#D5F4F7", "box": "#0A2D30", "accent": "#40E0D0"},
    "Karanlık Madde":       {"bg": "#020202", "text": "#808080", "box": "#0A0A0A", "accent": "#4B0082"},
    "Süpernova":            {"bg": "#1A0D00", "text": "#FFDAB9", "box": "#331A00", "accent": "#FF4500"},
    "Karadelik":            {"bg": "#000000", "text": "#4A4A4A", "box": "#050505", "accent": "#FFFFFF"},
    "Yıldız Tozu":          {"bg": "#0A0A12", "text": "#F0F8FF", "box": "#141424", "accent": "#FFD700"},
    "Kozmik Mavi":          {"bg": "#050B14", "text": "#B0C4DE", "box": "#0A1628", "accent": "#1E90FF"},
    "Güneş Patlaması":      {"bg": "#140500", "text": "#FFEFD5", "box": "#280A00", "accent": "#FF8C00"},
    "Zehirli Yeşil":        {"bg": "#051405", "text": "#E0FFE0", "box": "#0A280A", "accent": "#39FF14"},
    "Çikolata Rüyası":      {"bg": "#1E120D", "text": "#E8D8D0", "box": "#2D1B14", "accent": "#D2691E"},
    "Vanilya Gökyüzü":      {"bg": "#F3E5AB", "text": "#3E2723", "box": "#FFF3E0", "accent": "#5D4037"},
    "Kızıl Gezegen":        {"bg": "#1C0A00", "text": "#FFD1B3", "box": "#331400", "accent": "#B22222"},
    "Buz Devi":             {"bg": "#00111A", "text": "#CCEEFF", "box": "#002233", "accent": "#00FFFF"},
    "Volkanik Kül":         {"bg": "#1C1C1C", "text": "#A9A9A9", "box": "#2A2A2A", "accent": "#FF6347"},
    "Şafak Vakti":          {"bg": "#1A1016", "text": "#FADADD", "box": "#2B1A24", "accent": "#FF69B4"},
    "Alacakaranlık":        {"bg": "#0B0C10", "text": "#C5C6C7", "box": "#1F2833", "accent": "#66FCF1"},
    "Gece Kuşu":            {"bg": "#080808", "text": "#CCCCCC", "box": "#141414", "accent": "#7B68EE"},
    "Şehir Işıkları":       {"bg": "#0A0A0A", "text": "#F0F0F0", "box": "#171717", "accent": "#FF1493"},
    "Siber Gümüş":          {"bg": "#0D0D11", "text": "#D1D5DB", "box": "#1F2937", "accent": "#9CA3AF"},
    "Kripto Yeşili":        {"bg": "#0B110B", "text": "#E5FFE5", "box": "#152215", "accent": "#00FF00"},
    "Hacker Terminali":     {"bg": "#000000", "text": "#00FF00", "box": "#051105", "accent": "#008000"},
    "Galaktik Mor":         {"bg": "#0A0014", "text": "#E6CCFF", "box": "#140028", "accent": "#8A2BE2"},
    "Kuantum Foton":        {"bg": "#000A14", "text": "#CCE5FF", "box": "#001428", "accent": "#00BFFF"},
    "Biyolüminesans":       {"bg": "#001414", "text": "#CCFFFF", "box": "#002828", "accent": "#00FA9A"},
}

t_sec = tema_renkleri.get(tema, tema_renkleri["Galaksi (VIP)"])

st.markdown(f"""
    <style>
    @import url('{secili_font_url}');
    @import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@700&display=swap');
    .stApp {{ background-color: {t_sec['bg']}; color: {t_sec['text']}; font-family: '{secili_font}', sans-serif; }}
    [data-testid="stMetric"] {{ background: {t_sec['box']}; padding: 20px !important; border-radius: 12px !important; border: 1px solid {t_sec['accent']} !important; text-align: center; }}
    .kral-table {{ width: 100%; border-collapse: collapse; background: {t_sec['box']}22; margin-top: 10px; border: 1px solid {t_sec['accent']}33; border-radius: 10px; overflow: hidden; font-family: '{secili_font}', sans-serif; }}
    .kral-table th {{ padding: 12px; text-align: left; background: {t_sec['accent']}22; color: {t_sec['accent']}; font-weight: 700; border-bottom: 2px solid {t_sec['accent']}44; }}
    .kral-table td {{ padding: 12px; border-bottom: 1px solid {t_sec['accent']}11; color: {t_sec['text']}; }}
    .ticker-wrapper {{ width: 100%; overflow: hidden; background: {t_sec['box']}; border-radius: 8px; margin-bottom: 30px; padding: 15px 0; border: 1px solid {t_sec['accent']}44; }}
    .ticker-content {{ display: flex; animation: ticker 25s linear infinite; white-space: nowrap; gap: 60px; }}
    @keyframes ticker {{ 0% {{ transform: translateX(100%); }} 100% {{ transform: translateX(-100%); }} }}
    .up {{ color: #00e676; font-weight: bold; }} .down {{ color: #ff1744; font-weight: bold; }}
    .alarm-aktif {{ color: #ff1744; font-weight: bold; animation: blink 1s step-start infinite; }}
    @keyframes blink {{ 50% {{ opacity: 0; }} }}
    .indikator-bar {{ display: inline-block; height: 8px; border-radius: 4px; }}
    .vy-kart {{ background: {t_sec['box']}; border: 1px solid {t_sec['accent']}33; border-radius: 10px; padding: 12px 14px; margin-bottom: 8px; }}
    .vy-etiket {{ font-size: 10px; opacity: 0.5; margin-bottom: 2px; letter-spacing: 0.8px; }}
    .vy-deger {{ font-size: 13px; font-weight: 600; }}
    </style>
    """, unsafe_allow_html=True)

# ==========================================
# 2. ÜST BİLGİ VE PİYASA
# ==========================================
clock_html = f"""
<div style="position: fixed; top: 10px; right: 10px; background: {t_sec['box']}; padding: 10px 25px; border-radius: 15px; box-shadow: 0 4px 15px rgba(0,0,0,0.5); z-index: 99999; display: flex; flex-direction: column; align-items: flex-end; border: 1px solid {t_sec['accent']};">
    <div id="digital-clock" style="font-size: 20px; font-weight: bold; font-family: 'JetBrains Mono', monospace; color: {t_sec['accent']};"></div>
    <div id="date-display" style="font-size: 11px; font-weight: 600; color: {t_sec['text']}; opacity: 0.8; letter-spacing: 1px;"></div>
</div>
<script>
function updateClock() {{
    const trTime = new Date(new Date().toLocaleString("en-US", {{timeZone: "Europe/Istanbul"}}));
    document.getElementById('digital-clock').innerText = trTime.toLocaleTimeString('tr-TR', {{hour12: false}});
    document.getElementById('date-display').innerText = trTime.toLocaleDateString('tr-TR', {{day: '2-digit', month: 'long', year: 'numeric', weekday: 'long'}}).toUpperCase();
}}
setInterval(updateClock, 1000); updateClock();
</script>
"""
st.components.v1.html(clock_html, height=80)

st.markdown(f"<h2 style='text-align:center; color:{t_sec['accent']};'>🚀 Borsa Takip</h2>", unsafe_allow_html=True)

piyasa_izleme = {
    "BIST 100":  "XU100.IS",
    "USD/TRY":   "USDTRY=X",
    "EUR/TRY":   "EURTRY=X",
    "ONS ALTIN": "GC=F",
    "ONS GÜMÜŞ": "SI=F",
    "BTC":       "BTC-USD",
}
ticker_content = '<div class="ticker-wrapper"><div class="ticker-content">'
for isim, sembol in piyasa_izleme.items():
    d = fetch_stock_data(sembol)
    if d:
        try:
            last = float(d['hist']['Close'].iloc[-1])
            prev = float(d['hist']['Close'].iloc[-2])
            deg  = ((last - prev) / prev) * 100
            ticker_content += f'<div style="text-align:center;"><div>{isim}</div><div style="font-weight:bold;">{tr_format(last)}</div><div class="{"up" if deg>=0 else "down"}">{deg:+.2f}%</div></div>'
        except Exception as e:
            logger.warning(f"Ticker hatası ({sembol}): {e}")

# Gram Altın ve Gram Gümüş — önbellek sayesinde ek API çağrısı yok
try:
    _gc     = fetch_stock_data("GC=F")
    _si     = fetch_stock_data("SI=F")
    _usdtry = fetch_stock_data("USDTRY=X")
    if _gc and _usdtry:
        _ust  = float(_usdtry['hist']['Close'].iloc[-1])
        _gc_s = float(_gc['hist']['Close'].iloc[-1])
        _gc_p = float(_gc['hist']['Close'].iloc[-2])
        _gr_a = _gc_s * _ust / 31.1035
        _gr_ap = _gc_p * _ust / 31.1035
        _deg_a = ((_gr_a - _gr_ap) / _gr_ap) * 100
        ticker_content += (
            f'<div style="text-align:center;">'
            f'<div>GRAM ALTIN</div>'
            f'<div style="font-weight:bold;">{tr_format(_gr_a)}</div>'
            f'<div class="{"up" if _deg_a>=0 else "down"}">{_deg_a:+.2f}%</div>'
            f'</div>'
        )
    if _si and _usdtry:
        _ust  = float(_usdtry['hist']['Close'].iloc[-1])
        _si_s = float(_si['hist']['Close'].iloc[-1])
        _si_p = float(_si['hist']['Close'].iloc[-2])
        _gr_g = _si_s * _ust / 31.1035
        _gr_gp = _si_p * _ust / 31.1035
        _deg_g = ((_gr_g - _gr_gp) / _gr_gp) * 100
        ticker_content += (
            f'<div style="text-align:center;">'
            f'<div>GRAM GÜMÜŞ</div>'
            f'<div style="font-weight:bold;">{tr_format(_gr_g)}</div>'
            f'<div class="{"up" if _deg_g>=0 else "down"}">{_deg_g:+.2f}%</div>'
            f'</div>'
        )
except Exception as e:
    logger.warning(f"Gram fiyat hesaplama hatası: {e}")
st.markdown(ticker_content + '</div></div>', unsafe_allow_html=True)

# ==========================================
# 3. PARALEL VERİ HAZIRLAMA
# ==========================================
BIST_FULL = sorted([
    "A1CAP.IS","ACSEL.IS","ADEL.IS","ADESE.IS","AEFES.IS","AFYON.IS","AGESA.IS","AGHOL.IS",
    "AGROT.IS","AHGAZ.IS","AKBNK.IS","AKCNS.IS","AKENR.IS","AKFGY.IS","AKFYE.IS","AKGRT.IS",
    "AKMGY.IS","AKSA.IS","AKSEN.IS","ALARK.IS","ALBRK.IS","ALFAS.IS","ALGYO.IS","ALKIM.IS",
    "ALMAD.IS","ANELE.IS","ANGEN.IS","ANHYT.IS","ANSGR.IS","ARCLK.IS","ARDYZ.IS","ARENA.IS",
    "ARSAN.IS","ASGYO.IS","ASELS.IS","ASTOR.IS","ASUZU.IS","ATEKS.IS","ATLAS.IS","ATSYH.IS",
    "AVHOL.IS","AVOD.IS","AYDEM.IS","AYEN.IS","AYGAZ.IS","BAGFS.IS","BAKAB.IS","BALAT.IS",
    "BANVT.IS","BASGZ.IS","BAYRK.IS","BEGYO.IS","BERA.IS","BFREN.IS","BIMAS.IS","BINHO.IS",
    "BIOEN.IS","BIZIM.IS","BJKAS.IS","BLCYT.IS","BMSTL.IS","BNTAS.IS","BOBET.IS","BORLS.IS",
    "BORSK.IS","BOSSA.IS","BRISA.IS","BRKO.IS","BRKSN.IS","BRKVY.IS","BRLSM.IS","BRMEN.IS",
    "BRYAT.IS","BSOKE.IS","BTCIM.IS","BUCIM.IS","BURCE.IS","BURVA.IS","BVSAN.IS","BYDNR.IS",
    "CANTE.IS","CASA.IS","CATES.IS","CCOLA.IS","CELHA.IS","CEMAS.IS","CEMTS.IS","CEVNY.IS",
    "CIMSA.IS","CLEBI.IS","CMBTN.IS","CMENT.IS","CONSE.IS","COSMO.IS","CRDFA.IS","CRFSA.IS",
    "CUSAN.IS","CVKMD.IS","CWENE.IS","DAGHL.IS","DAGI.IS","DAPGM.IS","DARDL.IS","DENGE.IS",
    "DERAS.IS","DERIM.IS","DESA.IS","DESPC.IS","DEVA.IS","DGGYO.IS","DIRIT.IS","DITAS.IS",
    "DMSAS.IS","DOAS.IS","DOCO.IS","DOHOL.IS","DOKTA.IS","DURDO.IS","DYOBY.IS","DZGYO.IS",
    "EBEBK.IS","ECILC.IS","ECZYT.IS","EDATA.IS","EDIP.IS","EGEEN.IS","EGEPO.IS","EGGUB.IS",
    "EGPRO.IS","EGSER.IS","EKGYO.IS","EKIZ.IS","EKOS.IS","EKSUN.IS","ELITE.IS","EMKEL.IS",
    "ENERY.IS","ENJSA.IS","ENKAI.IS","ERBOS.IS","EREGL.IS","ERSU.IS","ESCOM.IS","ESEN.IS",
    "EUPWR.IS","EUREN.IS","EYGYO.IS","FMIZP.IS","FONET.IS","FORMT.IS","FORTE.IS","FROTO.IS",
    "FZLGY.IS","GARAN.IS","GENTS.IS","GEREL.IS","GESAN.IS","GIPTA.IS","GLBMD.IS","GLCVY.IS",
    "GLRYH.IS","GLYHO.IS","GMTAS.IS","GOKNR.IS","GOLTS.IS","GOODY.IS","GOZDE.IS","GRNYO.IS",
    "GRSEL.IS","GSDDE.IS","GSDHO.IS","GUBRF.IS","GWIND.IS","GZNMI.IS","HALKB.IS","HATEK.IS",
    "HATSN.IS","HEDEF.IS","HEKTS.IS","HKTM.IS","HLGYO.IS","HTTBT.IS","HUBVC.IS","HUNER.IS",
    "HURGZ.IS","ICBCT.IS","IDAS.IS","IDEAS.IS","IDGYO.IS","IEYHO.IS","IHEVA.IS","IHGZT.IS",
    "IHLAS.IS","IHLGM.IS","IHYAY.IS","IMASM.IS","INDES.IS","INFO.IS","INTEM.IS","IPEKE.IS",
    "ISATR.IS","ISBTR.IS","ISCTR.IS","ISDMR.IS","ISFIN.IS","ISGSY.IS","ISGYO.IS","ISMEN.IS",
    "ISSEN.IS","ISYAT.IS","ITTFH.IS","IZENR.IS","IZFAS.IS","IZINV.IS","IZMDC.IS","JANTS.IS",
    "KAPLM.IS","KAREL.IS","KARSN.IS","KARTN.IS","KARYE.IS","KATMR.IS","KAYSE.IS","KBCOR.IS",
    "KCAER.IS","KCHOL.IS","KFEIN.IS","KGYO.IS","KIMMR.IS","KLGYO.IS","KLMSN.IS","KLNMA.IS",
    "KLKIM.IS","KLRHO.IS","KLSYN.IS","KLYAS.IS","KMEPU.IS","KMPUR.IS","KNFRT.IS","KONTR.IS",
    "KONYA.IS","KORDS.IS","KOZAA.IS","KOZAL.IS","KRDMA.IS","KRDMB.IS","KRDMD.IS","KRGYO.IS",
    "KRONT.IS","KRPLS.IS","KRSTL.IS","KRTEK.IS","KRVGD.IS","KSTUR.IS","KUTPO.IS","KUVVA.IS",
    "KUYAS.IS","KZBGY.IS","KZGYO.IS","LIDER.IS","LIDFA.IS","LINK.IS","LMKDC.IS","LOGAS.IS",
    "LOGO.IS","LRSHO.IS","LUKSK.IS","MAALT.IS","MACKO.IS","MAGEN.IS","MAKIM.IS","MAKTK.IS",
    "MANAS.IS","MARKA.IS","MARTI.IS","MAVI.IS","MEDTR.IS","MEGAP.IS","MEKAG.IS","MEPET.IS",
    "MERCN.IS","MERKO.IS","METRO.IS","METUR.IS","MHRGY.IS","MIATK.IS","MIPAZ.IS","MNDRS.IS",
    "MNDTR.IS","MOBTL.IS","MPARK.IS","MRGYO.IS","MRSHL.IS","MSGYO.IS","MTRKS.IS","MUDO.IS",
    "MZHLD.IS","NATEN.IS","NETAS.IS","NIBAS.IS","NTGAZ.IS","NTHOL.IS","NUGYO.IS","NUHCM.IS",
    "OBAMS.IS","OBASE.IS","ODAS.IS","ONCSM.IS","ORCAY.IS","ORGE.IS","ORMA.IS","OSMEN.IS",
    "OSTIM.IS","OTKAR.IS","OYAKC.IS","OYAYO.IS","OYLUM.IS","OYYAT.IS","OZGYO.IS","OZKGY.IS",
    "OZRDN.IS","OZSUB.IS","PAGYO.IS","PAMEL.IS","PAPIL.IS","PARSN.IS","PASEU.IS","PATEK.IS",
    "PCILT.IS","PEGYO.IS","PEKGY.IS","PENTA.IS","PETKM.IS","PETUN.IS","PGSUS.IS","PINSU.IS",
    "PKART.IS","PKENT.IS","PNLSN.IS","PNSUT.IS","POLHO.IS","POLTK.IS","PRKAB.IS","PRKME.IS",
    "PRZMA.IS","PSDTC.IS","PSGYO.IS","QNBFB.IS","QNBFL.IS","QUAGR.IS","RALYH.IS","RAYYS.IS",
    "REEDR.IS","RNPOL.IS","RODRG.IS","ROYAL.IS","RTALB.IS","RUBNS.IS","RYGYO.IS","RYSAS.IS",
    "SAHOL.IS","SAMAT.IS","SANEL.IS","SANFO.IS","SANIC.IS","SARKY.IS","SASA.IS","SAYAS.IS",
    "SDTTR.IS","SEGYO.IS","SEKFK.IS","SEKOK.IS","SELEC.IS","SELGD.IS","SERVE.IS","SEYKM.IS",
    "SILVR.IS","SISE.IS","SKBNK.IS","SKTAS.IS","SKYMD.IS","SKYLP.IS","SMART.IS","SMRTG.IS",
    "SNGYO.IS","SNICA.IS","SNKPA.IS","SOKM.IS","SONME.IS","SRVGY.IS","SUMAS.IS","SUNTK.IS",
    "SURGY.IS","SUWEN.IS","TABGD.IS","TAPDI.IS","TARKM.IS","TATEN.IS","TATGD.IS","TAVHL.IS",
    "TBORG.IS","TCELL.IS","TDGYO.IS","TEKTU.IS","TERA.IS","TETMT.IS","TEZOL.IS","TGSAS.IS",
    "THYAO.IS","TIRE.IS","TKFEN.IS","TKNSA.IS","TMSN.IS","TOASO.IS","TRCAS.IS","TRGYO.IS",
    "TRILC.IS","TSKB.IS","TSPOR.IS","TTKOM.IS","TTRAK.IS","TUCLK.IS","TUKAS.IS","TUPRS.IS",
    "TURSG.IS","UFUK.IS","ULAS.IS","ULKER.IS","ULUFA.IS","ULUSE.IS","VAKBN.IS","VAKFN.IS",
    "VAKKO.IS","VANGD.IS","VBTYM.IS","VERTU.IS","VERUS.IS","VESBE.IS","VESTL.IS","VKGYO.IS",
    "VKING.IS","VRGYO.IS","YAPRK.IS","YATAS.IS","YAYLA.IS","YEOTK.IS","YESIL.IS","YGGYO.IS",
    "YGYO.IS","YKBNK.IS","YONGA.IS","YOTAS.IS","YUNSA.IS","YYLGD.IS","ZEDUR.IS","ZOREN.IS",
    "ZRGYO.IS",
])
FON_LIST = sorted([
    "TTE.IS","AES.IS","AFO.IS","AYA.IS","KPH.IS","KPA.IS","ZGD.IS","ZRE.IS",
    "TAU.IS","MAC.IS","YZG.IS","OPB.IS","NNF.IS","IDH.IS","GSP.IS","IHY.IS"
])

full_data = []
if st.session_state.portfoy:
    args_list = list(enumerate(st.session_state.portfoy))
    with ThreadPoolExecutor(max_workers=8) as executor:
        results = list(executor.map(fetch_single_item, args_list))
    full_data = sorted(results, key=lambda x: x['Hisse'])

# Günlük performans snapshot kaydet
kaydet_performans_snapshot(full_data)

# ==========================================
# ALARM KONTROLÜ — sayfa yüklenince otomatik çalışır
# ==========================================
tetiklenen_alarmlar = []
for alarm in st.session_state.alarmlar:
    hisse = alarm.get("Hisse", "")
    hedef = alarm.get("Hedef", 0.0)
    yon   = alarm.get("Yon", "Üstüne Çıkınca")
    aktif = alarm.get("Aktif", True)

    if not aktif:
        continue

    # Anlık fiyatı bul
    for row in full_data:
        if row["Hisse"] == hisse:
            guncel = row["Güncel"]
            if yon == "Üstüne Çıkınca" and guncel >= hedef:
                tetiklenen_alarmlar.append(f"🔔 {hisse}: {tr_format(guncel)} ₺ → Hedef {tr_format(hedef)} ₺ aşıldı!")
            elif yon == "Altına Düşünce" and guncel <= hedef:
                tetiklenen_alarmlar.append(f"🔔 {hisse}: {tr_format(guncel)} ₺ → Hedef {tr_format(hedef)} ₺ altına indi!")
            break

if tetiklenen_alarmlar:
    for msg in tetiklenen_alarmlar:
        st.warning(msg, icon="🚨")

# ==========================================
# 4. TABLAR VE İÇERİK
# ==========================================
tab_tr, tab_fon, tab_div, tab_ipo, tab_alarm, tab_analiz, tab_haberler, tab_temel, tab_notlar, tab_export = st.tabs([
    "🇹🇷 TÜRK BORSASI",
    "📊 YATIRIM FONLARI",
    "💰 TEMETTÜ GELİRİ",
    "🚀 HALKA ARZ TAKİP",
    "🔔 FİYAT ALARMLARI",
    "📈 ANALİZ",
    "📰 HABERLER",
    "🏦 TEMEL VERİLER",
    "📝 NOTLAR",
    "📤 DIŞA AKTAR",
])

with st.sidebar:
    # --- KARANLIK MOD TOGGLE ---
    _dm_col1, _dm_col2 = st.columns([3, 1])
    _dm_col1.markdown(
        f"<div style='padding-top:6px;font-size:12px;color:{t_sec['text']};opacity:0.7;'>"
        f"{'🌙 Karanlık Mod' if st.session_state.dark_mode else '☀️ Açık Mod'}</div>",
        unsafe_allow_html=True
    )
    if _dm_col2.button("⇄", key="dm_toggle", help="Karanlık/Açık mod"):
        st.session_state.dark_mode = not st.session_state.dark_mode
        # Tema'yı otomatik ayarla
        if st.session_state.dark_mode:
            if st.session_state.tema_secim in ["Siyah-Beyaz (Klasik)", "Vanilya Gökyüzü"]:
                st.session_state.tema_secim = "Galaksi (VIP)"
        else:
            st.session_state.tema_secim = "Siyah-Beyaz (Klasik)"
        st.rerun()

    # --- PİYASA DURUMU ---
    _acik, _durum_msg = piyasa_acik_mi()
    _durum_color = "#00e676" if _acik else "#ff1744"
    _durum_icon  = "🟢" if _acik else "🔴"
    st.markdown(
        f"<div style='background:{t_sec['box']};border:1px solid {_durum_color}55;"
        f"border-radius:10px;padding:8px 14px;margin-bottom:8px;"
        f"display:flex;justify-content:space-between;align-items:center;'>"
        f"<span style='font-size:11px;color:{t_sec['text']};opacity:0.6;'>BIST DURUMU</span>"
        f"<span style='font-size:11px;color:{_durum_color};font-weight:700;'>{_durum_icon} {_durum_msg}</span>"
        f"</div>",
        unsafe_allow_html=True
    )

    st.divider()

    # --- MİNİ PORTFÖY ÖZETİ ---
    if full_data:
        _deger   = sum(x['Değer']     for x in full_data)
        _kz      = sum(x['K/Z']       for x in full_data)
        _gunluk  = sum(x['DailyDiff'] for x in full_data)
        _pozisyon = len(full_data)
        _kz_color = "#00e676" if _kz     >= 0 else "#ff1744"
        _gd_color = "#00e676" if _gunluk >= 0 else "#ff1744"
        _box  = t_sec['box']
        _acc  = t_sec['accent']
        _txt  = t_sec['text']
        st.markdown(f"""
        <div style='background:{_box};border:1px solid {_acc}55;border-radius:12px;padding:14px 16px;margin-bottom:4px;'>
            <div style='color:{_acc};font-weight:700;font-size:11px;letter-spacing:1.5px;margin-bottom:10px;'>
                📊 PORTFÖY ÖZETİ
            </div>
            <div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:6px;'>
                <span style='color:{_txt};opacity:0.6;font-size:11px;'>Toplam Değer</span>
                <span style='color:{_txt};font-weight:600;font-size:12px;'>{tr_format(_deger)} ₺</span>
            </div>
            <div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:6px;'>
                <span style='color:{_txt};opacity:0.6;font-size:11px;'>Toplam K/Z</span>
                <span style='color:{_kz_color};font-weight:700;font-size:12px;'>{tr_format(_kz)} ₺</span>
            </div>
            <div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:6px;'>
                <span style='color:{_txt};opacity:0.6;font-size:11px;'>Günlük</span>
                <span style='color:{_gd_color};font-weight:700;font-size:12px;'>{tr_format(_gunluk)} ₺</span>
            </div>
            <div style='border-top:1px solid {_acc}22;margin-top:8px;padding-top:8px;display:flex;justify-content:space-between;'>
                <span style='color:{_txt};opacity:0.5;font-size:10px;'>Pozisyon</span>
                <span style='color:{_acc};font-size:10px;font-weight:600;'>{_pozisyon} adet</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

    st.divider()

    # --- YENİ VARLIK EKLE ---
    _acc2 = t_sec['accent']
    st.markdown(f"<div style='color:{_acc2};font-weight:700;font-size:13px;letter-spacing:1px;margin-bottom:8px;'>➕ YENİ VARLIK EKLE</div>", unsafe_allow_html=True)
    piyasa_sec = st.radio("Piyasa", ["Türk Borsası", "Yatırım Fonu"], horizontal=True)
    if piyasa_sec == "Türk Borsası":
        hisse_sec = st.selectbox("Hisse Seç", BIST_FULL)
    else:
        hisse_sec = st.selectbox("Fon Seç", FON_LIST + ["DİĞER"])
    if hisse_sec == "DİĞER":
        hisse_sec = st.text_input("Fon Kodu").upper()

    _col_a, _col_b = st.columns(2)
    adet_sec    = _col_a.number_input("Adet",    min_value=0,   step=1,  label_visibility="visible")
    maliyet_sec = _col_b.number_input("Maliyet (₺)", min_value=0.0, format="%.4f", label_visibility="visible")

    if st.button("🚀 Portföye Ekle", use_container_width=True):
        st.session_state.portfoy.append({
            "Piyasa": piyasa_sec, "Hisse": hisse_sec,
            "Adet": int(adet_sec), "Maliyet": float(maliyet_sec)
        })
        st.session_state.portfoy = sorted(st.session_state.portfoy, key=lambda x: x['Hisse'])
        save_json(PORTFOY_DOSYASI, st.session_state.portfoy)
        st.rerun()

    st.divider()

    # --- DÖVİZ ÇEVİRİCİ ---
    _acc3 = t_sec['accent']
    st.markdown(f"<div style='color:{_acc3};font-weight:700;font-size:13px;letter-spacing:1px;margin-bottom:8px;'>💱 DÖVİZ ÇEVİRİCİ</div>", unsafe_allow_html=True)
    try:
        _d_usd = fetch_stock_data("USDTRY=X")
        _d_eur = fetch_stock_data("EURTRY=X")
        _d_gc  = fetch_stock_data("GC=F")
        _d_si  = fetch_stock_data("SI=F")
        _usd_try_r = float(_d_usd['hist']['Close'].iloc[-1]) if _d_usd else 0.0
        _eur_try_r = float(_d_eur['hist']['Close'].iloc[-1]) if _d_eur else 0.0
        _gc_r      = float(_d_gc['hist']['Close'].iloc[-1])  if _d_gc  else 0.0
        _si_r      = float(_d_si['hist']['Close'].iloc[-1])  if _d_si  else 0.0
        _gram_altin = _gc_r * _usd_try_r / 31.1035 if _usd_try_r and _gc_r else 0.0
        _gram_gumus = _si_r * _usd_try_r / 31.1035 if _usd_try_r and _si_r else 0.0

        _doviz_rates = {
            "TRY (₺)":          1.0,
            "USD ($)":          _usd_try_r,
            "EUR (€)":          _eur_try_r,
            "Gram Altın":       _gram_altin,
            "Gram Gümüş":       _gram_gumus,
        }
        _cv1, _cv2 = st.columns(2)
        _cv_miktar = _cv1.number_input("Miktar", value=1.0, min_value=0.0, format="%.4f", key="cv_miktar")
        _cv_kaynak = _cv2.selectbox("Kaynak", list(_doviz_rates.keys()), key="cv_kaynak")
        _cv_hedef  = st.selectbox("Hedef", list(_doviz_rates.keys()), index=3, key="cv_hedef")

        if _doviz_rates[_cv_kaynak] > 0 and _doviz_rates[_cv_hedef] > 0:
            _sonuc = _cv_miktar * _doviz_rates[_cv_kaynak] / _doviz_rates[_cv_hedef]
            st.markdown(
                f"<div style='background:{t_sec['box']};border:1px solid {_acc3}44;"
                f"border-radius:8px;padding:10px 14px;text-align:center;'>"
                f"<span style='color:{t_sec['text']};opacity:0.6;font-size:11px;'>{tr_format(_cv_miktar)} {_cv_kaynak}</span><br>"
                f"<span style='color:{_acc3};font-size:18px;font-weight:700;'>{tr_format4(_sonuc)} {_cv_hedef}</span>"
                f"</div>",
                unsafe_allow_html=True
            )
    except Exception as _e:
        st.caption("Kur verisi yükleniyor...")

    st.divider()

    # --- KAR AL / ZARAR KES HESAPLAYICI ---
    st.markdown(f"<div style='color:{t_sec['accent']};font-weight:700;font-size:13px;letter-spacing:1px;margin-bottom:8px;'>🎯 KAR AL / ZARAR KES</div>", unsafe_allow_html=True)
    _portfoy_hisseler_kz = sorted(set(x['Hisse'] for x in full_data))
    if _portfoy_hisseler_kz:
        _kz_hisse = st.selectbox("Hisse", _portfoy_hisseler_kz, key="kz_hisse")
        _kz_guncel = next((x['Güncel'] for x in full_data if x['Hisse'] == _kz_hisse), 0.0)
    else:
        _kz_hisse  = None
        _kz_guncel = 0.0

    _kz_fiyat  = st.number_input("Güncel Fiyat (₺)", value=float(_kz_guncel), format="%.4f", key="kz_fiyat")
    _kz_c1, _kz_c2 = st.columns(2)
    _kar_pct   = _kz_c1.number_input("Kar Al %", value=10.0, min_value=0.1, step=0.5, key="kar_pct")
    _zarar_pct = _kz_c2.number_input("Zarar Kes %", value=5.0, min_value=0.1, step=0.5, key="zarar_pct")

    if _kz_fiyat > 0:
        _kar_fiyat   = _kz_fiyat * (1 + _kar_pct / 100)
        _zarar_fiyat = _kz_fiyat * (1 - _zarar_pct / 100)
        st.markdown(
            f"<div style='background:{t_sec['box']};border:1px solid {t_sec['accent']}33;"
            f"border-radius:8px;padding:10px 14px;margin-top:4px;'>"
            f"<div style='display:flex;justify-content:space-between;margin-bottom:6px;'>"
            f"  <span style='font-size:11px;opacity:0.55;'>Kar Al Hedefi</span>"
            f"  <span style='color:#00e676;font-weight:700;'>{tr_format4(_kar_fiyat)} ₺</span>"
            f"</div>"
            f"<div style='display:flex;justify-content:space-between;'>"
            f"  <span style='font-size:11px;opacity:0.55;'>Zarar Kes Seviyesi</span>"
            f"  <span style='color:#ff1744;font-weight:700;'>{tr_format4(_zarar_fiyat)} ₺</span>"
            f"</div>"
            f"</div>",
            unsafe_allow_html=True
        )

    st.divider()

    # --- VERİ KORUMA PANELİ ---
    _acc_v = t_sec['accent']
    _txt_v = t_sec['text']
    _box_v = t_sec['box']

    # Kaç hisse var ve JSON nerede?
    _hisse_sayisi = len(st.session_state.portfoy)
    _json_var     = os.path.exists(PORTFOY_DOSYASI)
    _bak_var      = os.path.exists(PORTFOY_DOSYASI + ".bak")
    _json_boyut   = round(os.path.getsize(PORTFOY_DOSYASI) / 1024, 1) if _json_var else 0
    _durum_renk   = "#00e676" if _json_var and _hisse_sayisi > 0 else "#ff1744"

    st.markdown(
        f"<div style='background:{_box_v};border:1px solid {_acc_v}33;"
        f"border-radius:10px;padding:10px 14px;'>"
        f"<div style='color:{_acc_v};font-weight:700;font-size:11px;letter-spacing:1px;margin-bottom:8px;'>"
        f"💾 VERİ KORUMA"
        f"</div>"
        f"<div style='display:flex;justify-content:space-between;margin-bottom:4px;'>"
        f"<span style='font-size:10px;opacity:0.55;'>Kayıtlı hisse</span>"
        f"<span style='color:{_durum_renk};font-size:10px;font-weight:700;'>{_hisse_sayisi} adet</span>"
        f"</div>"
        f"<div style='display:flex;justify-content:space-between;margin-bottom:4px;'>"
        f"<span style='font-size:10px;opacity:0.55;'>JSON dosyası</span>"
        f"<span style='font-size:10px;color:{_durum_renk};'>{'✅ Var · ' + str(_json_boyut) + ' KB' if _json_var else '❌ Bulunamadı'}</span>"
        f"</div>"
        f"<div style='display:flex;justify-content:space-between;'>"
        f"<span style='font-size:10px;opacity:0.55;'>Yedek (.bak)</span>"
        f"<span style='font-size:10px;color:{'#00e676' if _bak_var else '#888'};'>{'✅ Var' if _bak_var else '—'}</span>"
        f"</div>"
        f"<div style='font-size:9px;opacity:0.35;margin-top:6px;word-break:break-all;'>"
        f"{os.path.dirname(PORTFOY_DOSYASI)}"
        f"</div>"
        f"</div>",
        unsafe_allow_html=True
    )

    # El ile yedek al butonu
    if st.button("📁 Şimdi Yedek Al", key="manuel_yedek", use_container_width=True):
        try:
            import shutil, zipfile
            _zip_adi = os.path.join(_VERI_DIZIN,
                f"portfoy_yedek_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip")
            with zipfile.ZipFile(_zip_adi, 'w', zipfile.ZIP_DEFLATED) as zf:
                for _dosya in [PORTFOY_DOSYASI, IPO_DOSYASI, ALARM_DOSYASI,
                               PERFORMANS_DOSYASI, NOTLAR_DOSYASI]:
                    if os.path.exists(_dosya):
                        zf.write(_dosya, os.path.basename(_dosya))
            st.success(f"✅ Yedek oluşturuldu: {os.path.basename(_zip_adi)}")
        except Exception as _ez:
            st.error(f"Yedek alınamadı: {_ez}")

    st.divider()

    # --- YENİLEME & TELEFON SYNC ---
    _acc_s = t_sec['accent']
    st.markdown(
        f"<div style='color:{_acc_s};font-weight:700;font-size:11px;"
        f"letter-spacing:1px;margin-bottom:6px;'>🔄 YENİLEME & SYNC</div>",
        unsafe_allow_html=True
    )
    _sure_sec = st.select_slider(
        "Otomatik yenileme",
        options=[15, 30, 60, 120, 300],
        value=st.session_state.yenileme_suresi,
        format_func=lambda x: f"{x}sn" if x < 60 else f"{x//60}dk",
        key="yenileme_slider"
    )
    if _sure_sec != st.session_state.yenileme_suresi:
        st.session_state.yenileme_suresi = _sure_sec
        st.rerun()

    # Telefon URL bilgisi
    try:
        import socket
        _hostname = socket.gethostname()
        _local_ip = socket.gethostbyname(_hostname)
    except Exception:
        _local_ip = "—"

    st.markdown(
        f"<div style='background:{t_sec['box']};border:1px solid {_acc_s}22;"
        f"border-radius:8px;padding:8px 12px;margin-top:4px;'>"
        f"<div style='font-size:10px;opacity:0.5;margin-bottom:4px;'>📱 TELEFONDA AÇMAK İÇİN</div>"
        f"<div style='font-size:10px;color:{_acc_s};font-weight:600;'>"
        f"Aynı Wi-Fi'ya bağlı ol,<br>tarayıcıda şunu aç:</div>"
        f"<div style='font-size:11px;font-family:monospace;margin-top:4px;"
        f"word-break:break-all;opacity:0.8;'>"
        f"http://{_local_ip}:8501</div>"
        f"<div style='font-size:9px;opacity:0.4;margin-top:4px;'>"
        f"Aynı URL'yi telefon ve bilgisayarda aç — "
        f"her ikisi de {st.session_state.yenileme_suresi}sn'de otomatik güncellenir.</div>"
        f"</div>",
        unsafe_allow_html=True
    )


    acc = t_sec['accent']
    box = t_sec['box']
    txt = t_sec['text']
    with st.expander("🛠️ VARLIK YÖNETİMİ"):
        for r in df_local.iterrows():
            kz_color = "#00e676" if r['K/Z'] >= 0 else "#ff1744"
            kz_pct   = ((r['Güncel'] - r['Maliyet']) / r['Maliyet'] * 100) if r['Maliyet'] > 0 else 0.0
            sinyal_str = str(r.get('Sinyal', '—'))

            # Bilgi kartı
            st.markdown(
                f"<div class='vy-kart'>"
                f"<div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;'>"
                f"  <span style='color:{acc};font-weight:700;font-size:15px;letter-spacing:0.5px;'>{r['Hisse']}</span>"
                f"  <span style='color:{txt};opacity:0.45;font-size:10px;letter-spacing:1px;'>{r['Piyasa'].upper()}</span>"
                f"</div>"
                f"<div style='display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:10px;'>"
                f"  <div><div class='vy-etiket'>ADET</div>"
                f"      <div class='vy-deger' style='color:{txt};'>{r['Adet']}</div></div>"
                f"  <div><div class='vy-etiket'>MALİYET</div>"
                f"      <div class='vy-deger' style='color:{txt};'>{tr_format4(r['Maliyet'])} ₺</div></div>"
                f"  <div><div class='vy-etiket'>GÜNCEL</div>"
                f"      <div class='vy-deger' style='color:{acc};'>{tr_format4(r['Güncel'])} ₺</div></div>"
                f"  <div><div class='vy-etiket'>K/Z</div>"
                f"      <div class='vy-deger' style='color:{kz_color};'>{tr_format(r['K/Z'])} ₺"
                f"          <span style='font-size:10px;opacity:0.8;'> ({kz_pct:+.1f}%)</span></div></div>"
                f"</div>"
                f"<div style='margin-top:8px;padding-top:8px;border-top:1px solid {acc}18;"
                f"display:flex;justify-content:space-between;align-items:center;'>"
                f"  <span style='font-size:11px;opacity:0.45;'>Sinyal: {sinyal_str}</span>"
                f"  <span style='font-size:11px;opacity:0.45;'>Toplam: {tr_format(r['Değer'])} ₺</span>"
                f"</div>"
                f"</div>",
                unsafe_allow_html=True
            )

            # Düzenleme satırı
            ec1, ec2, ec3, ec4 = st.columns([2, 2, 1, 1])
            y_adet    = ec1.number_input("Yeni Adet",    value=int(r['Adet']),      step=1,        key=f"a_{r['id']}")
            y_maliyet = ec2.number_input("Yeni Maliyet (₺)", value=float(r['Maliyet']), format="%.4f", key=f"m_{r['id']}")
            ec3.markdown("<div style='margin-top:28px;'></div>", unsafe_allow_html=True)
            ec4.markdown("<div style='margin-top:28px;'></div>", unsafe_allow_html=True)

            if ec3.button("💾 Kaydet", key=f"s_{r['id']}", use_container_width=True):
                st.session_state.portfoy[r['id']]['Adet']    = y_adet
                st.session_state.portfoy[r['id']]['Maliyet'] = y_maliyet
                st.session_state.portfoy = sorted(st.session_state.portfoy, key=lambda x: x['Hisse'])
                save_json(PORTFOY_DOSYASI, st.session_state.portfoy)
                st.rerun()
            if ec4.button("❌ Sil", key=f"d_{r['id']}", use_container_width=True):
                st.session_state.portfoy.pop(r['id'])
                save_json(PORTFOY_DOSYASI, st.session_state.portfoy)
                st.rerun()
            st.markdown("<div style='margin-bottom:4px;'></div>", unsafe_allow_html=True)

def make_sparkline_svg(prices, width=80, height=28, renk_kz=None):
    """
    Verilen fiyat listesinden inline SVG sparkline üretir.
    renk_kz: None ise ilk-son fiyata göre otomatik renk seçer.
    """
    if not prices or len(prices) < 2:
        return "<span style='opacity:0.25;font-size:10px;'>—</span>"

    try:
        mn   = min(prices)
        mx   = max(prices)
        span = mx - mn if mx != mn else 1.0
        pad  = 3   # üst/alt boşluk (px)

        # x koordinatları eşit aralıklı
        xs = [round(i * (width - 1) / (len(prices) - 1), 2) for i in range(len(prices))]
        # y koordinatları: yüksek fiyat = düşük y (SVG y ekseni ters)
        ys = [round(pad + (1 - (p - mn) / span) * (height - 2 * pad), 2) for p in prices]

        # Renk: son - ilk fiyata göre
        if renk_kz is not None:
            renk = "#00e676" if renk_kz >= 0 else "#ff1744"
        else:
            renk = "#00e676" if prices[-1] >= prices[0] else "#ff1744"

        # Polylon noktaları
        pts = " ".join(f"{x},{y}" for x, y in zip(xs, ys))

        # Dolgu için alan kapatma (altına in, sola git, kapat)
        alan_pts = (
            f"0,{height} "      # sol alt köşe
            + pts +
            f" {width},{height}" # sağ alt köşe
        )

        svg = (
            f"<svg width='{width}' height='{height}' viewBox='0 0 {width} {height}' "
            f"xmlns='http://www.w3.org/2000/svg' style='vertical-align:middle;overflow:visible;'>"
            # Alan dolgusu (şeffaf)
            f"<polygon points='{alan_pts}' fill='{renk}' fill-opacity='0.12'/>"
            # Çizgi
            f"<polyline points='{pts}' fill='none' stroke='{renk}' "
            f"stroke-width='1.5' stroke-linecap='round' stroke-linejoin='round'/>"
            # Son nokta
            f"<circle cx='{xs[-1]}' cy='{ys[-1]}' r='2' fill='{renk}'/>"
            f"</svg>"
        )
        return svg
    except Exception:
        return "<span style='opacity:0.25;font-size:10px;'>—</span>"


@st.cache_data(ttl=600)
def hesapla_korelasyon(hisse_listesi):
    """Son 30 gün kapanış fiyatları üzerinden korelasyon matrisi hesaplar."""
    kapanis = {}
    for h in hisse_listesi:
        d = fetch_stock_data(h)
        if d and not d['hist'].empty:
            kapanis[h] = d['hist']['Close'].values[-30:]
    if len(kapanis) < 2:
        return None
    min_uzunluk = min(len(v) for v in kapanis.values())
    df_kap = pd.DataFrame({k: v[-min_uzunluk:] for k, v in kapanis.items()})
    return df_kap.corr()

# ==========================================
# GELİŞTİRİLMİŞ TABLO — RSI + MACD + BB sütunu
# ==========================================
def render_kral_table(df_local, goster_indikatör=True):
    if goster_indikatör:
        baslik = "<tr><th>VARLIK</th><th>7G</th><th>SİNYAL</th><th>RSI</th><th>MACD-H</th><th>BB%</th><th>ADET</th><th>MALİYET</th><th>GÜNCEL</th><th>K/Z</th><th>TOPLAM</th></tr>"
    else:
        baslik = "<tr><th>VARLIK</th><th>7G</th><th>SİNYAL</th><th>ADET</th><th>MALİYET</th><th>GÜNCEL</th><th>K/Z</th><th>TOPLAM</th></tr>"

    table_html = f"<table class='kral-table'><thead>{baslik}</thead><tbody>"

    for _, r in df_local.iterrows():
        kz_color = "#00e676" if r['K/Z'] >= 0 else "#ff1744"

        # Sparkline SVG
        spark_prices = r.get('Sparkline', [])
        spark_svg    = make_sparkline_svg(spark_prices, renk_kz=r['K/Z'])

        if goster_indikatör:
            # RSI rengi
            rsi = r['RSI']
            if rsi < 35:   rsi_color = "#00e676"
            elif rsi > 65: rsi_color = "#ff1744"
            else:          rsi_color = "#ffc107"

            # MACD histogram
            macd_h = r['MACD_H']
            macd_color  = "#00e676" if macd_h > 0 else "#ff1744"
            macd_sembol = f"+{tr_format(macd_h)}" if macd_h > 0 else tr_format(macd_h)

            # Bollinger % bar
            bb_pct   = max(0, min(100, r['BB_PCT']))
            bb_color = "#00e676" if bb_pct < 30 else ("#ff1744" if bb_pct > 70 else "#ffc107")
            bb_bg    = t_sec['box']

            extra = (
                f"<td style='color:{rsi_color};font-weight:bold;'>{rsi:.1f}</td>"
                f"<td style='color:{macd_color};font-weight:bold;'>{macd_sembol}</td>"
                f"<td>"
                f"<div style='background:{bb_bg}; border-radius:4px; width:80px; height:8px; display:inline-block; vertical-align:middle;'>"
                f"<div style='width:{bb_pct}%; background:{bb_color}; height:8px; border-radius:4px;'></div>"
                f"</div> <span style='font-size:11px;color:{bb_color};'>{bb_pct:.0f}%</span>"
                f"</td>"
                f"<td>{r['Adet']}</td>"
            )
            table_html += (
                f"<tr>"
                f"<td><b>{r['Hisse']}</b></td>"
                f"<td style='padding:8px 12px;'>{spark_svg}</td>"
                f"<td>{r['Sinyal']}</td>"
                f"{extra}"
                f"<td>{tr_format4(r['Maliyet'])} ₺</td>"
                f"<td>{tr_format4(r['Güncel'])} ₺</td>"
                f"<td style='color:{kz_color};font-weight:bold;'>{tr_format(r['K/Z'])} ₺</td>"
                f"<td><b>{tr_format(r['Değer'])} ₺</b></td>"
                f"</tr>"
            )
        else:
            table_html += (
                f"<tr>"
                f"<td><b>{r['Hisse']}</b></td>"
                f"<td style='padding:8px 12px;'>{spark_svg}</td>"
                f"<td>{r['Sinyal']}</td>"
                f"<td>{r['Adet']}</td>"
                f"<td>{tr_format4(r['Maliyet'])} ₺</td>"
                f"<td>{tr_format4(r['Güncel'])} ₺</td>"
                f"<td style='color:{kz_color};font-weight:bold;'>{tr_format(r['K/Z'])} ₺</td>"
                f"<td><b>{tr_format(r['Değer'])} ₺</b></td>"
                f"</tr>"
            )

    return table_html + "</tbody></table>"

# ==========================================
# TÜRK BORSASI TABU
# ==========================================
with tab_tr:
    df_bist = pd.DataFrame([x for x in full_data if x['Piyasa'] == 'Türk Borsası'])
    if not df_bist.empty:
        m1, m2, m3 = st.columns(3)
        m1.metric("PORTFÖY DEĞERİ",  f"{tr_format(df_bist['Değer'].sum())} ₺")
        m2.metric("TOPLAM K/Z",       f"{tr_format(df_bist['K/Z'].sum())} ₺")
        m3.metric("GÜNLÜK DEĞİŞİM",  f"{tr_format(df_bist['DailyDiff'].sum())} ₺")

        st.markdown(
            f"<small style='color:{t_sec['accent']}88;'>RSI: &lt;35 aşırı satım &gt;65 aşırı alım | MACD-H: + bullish / - bearish | BB%: &lt;30 alt bant &gt;70 üst bant</small>",
            unsafe_allow_html=True
        )
        st.markdown(render_kral_table(df_bist, goster_indikatör=True), unsafe_allow_html=True)
        varlik_yonetimi_render(df_bist)

        df_chart = df_bist[df_bist['Değer'] > 0]
        if not df_chart.empty:
            st.divider()
            _total  = df_chart['Değer'].sum()
            _labels = df_chart['Hisse'].tolist()
            _values = df_chart['Değer'].tolist()
            _palette = [
                "#00D4FF","#FF6B6B","#FFD93D","#6BCB77","#A78BFA",
                "#F97316","#38BDF8","#F472B6","#34D399","#FBBF24",
                "#E879F9","#60A5FA","#FB923C","#4ADE80","#C084FC",
            ]
            _colors = (_palette * (len(_labels) // len(_palette) + 1))[:len(_labels)]

            fig = go.Figure(data=[go.Pie(
                labels=_labels,
                values=_values,
                hole=0.58,
                textinfo='label+percent',
                textposition='outside',
                hovertemplate='<b>%{label}</b><br>Değer: %{value:,.2f} ₺<br>Oran: %{percent}<extra></extra>',
                marker=dict(
                    colors=_colors,
                    line=dict(color=t_sec['bg'], width=2)
                ),
                pull=[0.02] * len(_labels),
                domain=dict(x=[0.0, 0.72], y=[0.0, 1.0]),
            )])
            fig.update_layout(
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color=t_sec['text'], size=12, family=secili_font),
                showlegend=True,
                legend=dict(
                    orientation='v',
                    yanchor='middle', y=0.5,
                    xanchor='left',   x=0.75,
                    font=dict(size=11, color=t_sec['text']),
                    bgcolor='rgba(0,0,0,0)',
                    itemsizing='constant',
                ),
                margin=dict(t=70, b=30, l=20, r=20),
                height=480,
                title=dict(
                    text="📊 Hisse Dağılımı",
                    font=dict(size=15, color=t_sec['accent']),
                    x=0.36, xanchor='center', y=0.97,
                ),
                annotations=[dict(
                    text=f"<b>{tr_format(_total)}</b><br>₺ TOPLAM",
                    x=0.36, y=0.5,
                    font=dict(size=14, color=t_sec['accent']),
                    showarrow=False, align='center',
                    xref='paper', yref='paper',
                )]
            )
            st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("Hisse senedi bulunamadı.")

# ==========================================
# YATIRIM FONLARI TABU
# ==========================================
with tab_fon:
    df_fon = pd.DataFrame([x for x in full_data if x['Piyasa'] == 'Yatırım Fonu'])
    if not df_fon.empty:
        # TEFAS hata uyarıları
        hata_fonlar = df_fon[df_fon['Sinyal'].str.startswith('⚠️', na=False)]
        if not hata_fonlar.empty:
            st.warning(
                f"⚠️ {', '.join(hata_fonlar['Hisse'].tolist())} için TEFAS verisi alınamadı. "
                "Maliyet fiyatı gösteriliyor.",
                icon="⚠️"
            )

        mf1, mf2, mf3 = st.columns(3)
        mf1.metric("FON PORTFÖY DEĞERİ", f"{tr_format(df_fon['Değer'].sum())} ₺")
        mf2.metric("TOPLAM K/Z",          f"{tr_format(df_fon['K/Z'].sum())} ₺")
        mf3.metric("GÜNLÜK DEĞİŞİM",     f"{tr_format(df_fon['DailyDiff'].sum())} ₺")

        st.markdown(render_kral_table(df_fon, goster_indikatör=False), unsafe_allow_html=True)
        varlik_yonetimi_render(df_fon)

        df_chart_f = df_fon[df_fon['Değer'] > 0]
        if not df_chart_f.empty:
            st.divider()
            _total_f  = df_chart_f['Değer'].sum()
            _labels_f = df_chart_f['Hisse'].tolist()
            _values_f = df_chart_f['Değer'].tolist()
            _palette_f = [
                "#A78BFA","#34D399","#F97316","#38BDF8","#F472B6",
                "#FBBF24","#60A5FA","#6BCB77","#E879F9","#00D4FF",
                "#FB923C","#4ADE80","#C084FC","#FFD93D","#FF6B6B",
            ]
            _colors_f = (_palette_f * (len(_labels_f) // len(_palette_f) + 1))[:len(_labels_f)]

            fig_f = go.Figure(data=[go.Pie(
                labels=_labels_f,
                values=_values_f,
                hole=0.58,
                textinfo='label+percent',
                textposition='outside',
                hovertemplate='<b>%{label}</b><br>Değer: %{value:,.2f} ₺<br>Oran: %{percent}<extra></extra>',
                marker=dict(
                    colors=_colors_f,
                    line=dict(color=t_sec['bg'], width=2)
                ),
                pull=[0.02] * len(_labels_f),
                domain=dict(x=[0.0, 0.72], y=[0.0, 1.0]),
            )])
            fig_f.update_layout(
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color=t_sec['text'], size=12, family=secili_font),
                showlegend=True,
                legend=dict(
                    orientation='v',
                    yanchor='middle', y=0.5,
                    xanchor='left',   x=0.75,
                    font=dict(size=11, color=t_sec['text']),
                    bgcolor='rgba(0,0,0,0)',
                    itemsizing='constant',
                ),
                margin=dict(t=70, b=30, l=20, r=20),
                height=480,
                title=dict(
                    text="📊 Fon Dağılımı",
                    font=dict(size=15, color=t_sec['accent']),
                    x=0.36, xanchor='center', y=0.97,
                ),
                annotations=[dict(
                    text=f"<b>{tr_format(_total_f)}</b><br>₺ TOPLAM",
                    x=0.36, y=0.5,
                    font=dict(size=14, color=t_sec['accent']),
                    showarrow=False, align='center',
                    xref='paper', yref='paper',
                )]
            )
            st.plotly_chart(fig_f, use_container_width=True)
    else:
        st.info("Fon bulunamadı.")

# ==========================================
# TEMETTÜ GELİRİ TABU — Düzeltilmiş
# ==========================================
with tab_div:
    st.markdown("### 💰 Yıllık Projeksiyon (Net Değerler)")
    st.markdown(
        f"<small style='color:{t_sec['accent']}88;'>* işareti: Son 1 yılda dağıtım yapılmamış, gösterilen değer en son temettüdür.</small>",
        unsafe_allow_html=True
    )

    df_div = pd.DataFrame([x for x in full_data if x['Temettu'] > 0])
    if not df_div.empty:
        toplam_temettu = df_div['NetTemettu'].sum()
        st.metric(
            "TAHMİNİ YILLIK NAKİT AKIŞI",
            f"{tr_format(toplam_temettu)} ₺",
            delta=f"Aylık: {tr_format(toplam_temettu / 12)} ₺"
        )

        div_table = (
            "<table class='kral-table'><thead>"
            "<tr>"
            "<th>HİSSE</th>"
            "<th>SON DAĞITIM TARİHİ</th>"
            "<th>ADET</th>"
            "<th>BRÜT HİSSE BAŞI</th>"
            "<th>NET HİSSE BAŞI (%10 stopaj)</th>"
            "<th>YILLIK TOPLAM (NET)</th>"
            "<th>NET VERİM (%)</th>"
            "</tr>"
            "</thead><tbody>"
        )
        for _, r in df_div.iterrows():
            verim = (r['Temettu'] / r['Güncel']) * 100 if r['Güncel'] > 0 else 0
            brut  = round(r['Temettu'] / 0.90, 4)   # Net'ten brüte çevir
            tarih_goster = r['Tarih']
            tarih_color  = t_sec['text'] if '*' not in str(tarih_goster) else "#ffc107"

            div_table += (
                f"<tr>"
                f"<td><b>{r['Hisse']}</b></td>"
                f"<td style='color:{tarih_color};'>{tarih_goster}</td>"
                f"<td>{r['Adet']}</td>"
                f"<td style='color:#ffc107;'>{tr_format4(brut)} ₺</td>"
                f"<td style='color:#00e676;'>{tr_format4(r['Temettu'])} ₺</td>"
                f"<td><b style='color:#00e676;'>{tr_format(r['NetTemettu'])} ₺</b></td>"
                f"<td>%{verim:.2f}</td>"
                f"</tr>"
            )
        st.markdown(div_table + "</tbody></table>", unsafe_allow_html=True)
    else:
        st.warning("Temettü verisi bulunamadı.")

# ==========================================
# HALKA ARZ TABU
# ==========================================
with tab_ipo:
    st.subheader("🚀 Yeni Halka Arz Ekle")
    with st.form("ipo_form", clear_on_submit=True):
        ic1, ic2, ic3 = st.columns(3)
        ipo_isim  = ic1.text_input("Şirket Kodu (Örn: BINHO)")
        ipo_fiyat = ic2.number_input("Halka Arz Fiyatı", min_value=0.0)
        ipo_adet  = ic3.number_input("Lot Sayısı", min_value=0, step=1)
        if st.form_submit_button("➕ Listeye Ekle"):
            if ipo_isim:
                st.session_state.ipo_liste.append({
                    "Isim": ipo_isim.upper(),
                    "Fiyat": ipo_fiyat,
                    "Adet": int(ipo_adet)
                })
                save_json(IPO_DOSYASI, st.session_state.ipo_liste)
                st.rerun()

    if st.session_state.ipo_liste:
        for idx, ipo in enumerate(st.session_state.ipo_liste):
            maliyet = ipo['Adet'] * ipo['Fiyat']
            baslik  = f"📈 {ipo['Isim']}  |  {tr_format(ipo['Fiyat'])} ₺  |  {ipo['Adet']} Lot"
            with st.expander(baslik):
                # Özet satır → lot, fiyat, maliyet, SİL butonu hizalı
                ri1, ri2, ri3, ri4 = st.columns([2, 2, 2, 1])
                ri1.markdown(
                    f"<div style='padding:8px 0;'>"
                    f"<div style='font-size:10px;opacity:0.55;'>HALKA ARZ FİYATI</div>"
                    f"<div style='font-weight:700;font-size:14px;color:{t_sec['accent']};'>{tr_format(ipo['Fiyat'])} ₺</div>"
                    f"</div>",
                    unsafe_allow_html=True
                )
                ri2.markdown(
                    f"<div style='padding:8px 0;'>"
                    f"<div style='font-size:10px;opacity:0.55;'>LOT SAYISI</div>"
                    f"<div style='font-weight:700;font-size:14px;'>{ipo['Adet']} Lot</div>"
                    f"</div>",
                    unsafe_allow_html=True
                )
                ri3.markdown(
                    f"<div style='padding:8px 0;'>"
                    f"<div style='font-size:10px;opacity:0.55;'>TOPLAM MALİYET</div>"
                    f"<div style='font-weight:700;font-size:14px;'>{tr_format(maliyet)} ₺</div>"
                    f"</div>",
                    unsafe_allow_html=True
                )
                ri4.markdown("<div style='padding-top:18px;'></div>", unsafe_allow_html=True)
                if ri4.button("❌ SİL", key=f"del_ipo_{idx}", use_container_width=True):
                    st.session_state.ipo_liste.pop(idx)
                    save_json(IPO_DOSYASI, st.session_state.ipo_liste)
                    st.rerun()

                st.divider()

                # Tavan simülasyon tablosu
                tavan_html = (
                    "<table class='kral-table' style='text-align:center;'>"
                    "<thead><tr>"
                    "<th style='text-align:center;'>GÜN</th>"
                    "<th style='text-align:center;'>FİYAT</th>"
                    "<th style='text-align:center;'>TOPLAM KAR</th>"
                    "</tr></thead><tbody>"
                )
                p = ipo['Fiyat']
                for g in range(1, 11):
                    p  *= 1.10
                    kar = (p * ipo['Adet']) - maliyet
                    tavan_html += (
                        f"<tr><td><b>{g}. Tavan</b></td>"
                        f"<td>{tr_format(p)} ₺</td>"
                        f"<td style='color:#00e676;font-weight:bold;'>+{tr_format(kar)} ₺</td></tr>"
                    )
                tavan_html += "</tbody></table>"
                st.markdown(tavan_html, unsafe_allow_html=True)

                # Portföye Aktar butonu
                st.markdown("<div style='margin-top:12px;'></div>", unsafe_allow_html=True)
                pa1, pa2, pa3 = st.columns([2, 2, 1])
                aktar_hisse = pa1.text_input(
                    "BIST Kodu (.IS ekli)",
                    value=ipo['Isim'] + ".IS",
                    key=f"aktar_kod_{idx}",
                    placeholder="Örn: BINHO.IS"
                )
                aktar_maliyet = pa2.number_input(
                    "Alış Maliyeti (₺)",
                    value=float(ipo['Fiyat']),
                    format="%.4f",
                    key=f"aktar_maliyet_{idx}"
                )
                pa3.markdown("<div style='margin-top:28px;'></div>", unsafe_allow_html=True)
                if pa3.button("📥 Portföye Ekle", key=f"aktar_btn_{idx}", use_container_width=True):
                    if aktar_hisse:
                        st.session_state.portfoy.append({
                            "Piyasa": "Türk Borsası",
                            "Hisse":  aktar_hisse.upper(),
                            "Adet":   int(ipo['Adet']),
                            "Maliyet": float(aktar_maliyet)
                        })
                        st.session_state.portfoy = sorted(
                            st.session_state.portfoy, key=lambda x: x['Hisse']
                        )
                        save_json(PORTFOY_DOSYASI, st.session_state.portfoy)
                        st.success(f"✅ {aktar_hisse.upper()} portföye eklendi!")
                        st.rerun()

# ==========================================
# FİYAT ALARMLARI TABU — YENİ
# ==========================================
with tab_alarm:
    st.subheader("🔔 Fiyat Alarmı Ekle")
    st.markdown(
        f"<small style='color:{t_sec['accent']}88;'>"
        "Belirlediğin fiyat seviyesine ulaşıldığında sayfa başında uyarı gösterilir."
        "</small>",
        unsafe_allow_html=True
    )

    # Portföydeki tüm hisseler + manuel giriş
    portfoy_hisseler = sorted(set(x['Hisse'] for x in st.session_state.portfoy))

    with st.form("alarm_form", clear_on_submit=True):
        al1, al2, al3, al4 = st.columns([2, 2, 2, 1])
        if portfoy_hisseler:
            alarm_hisse = al1.selectbox("Hisse", portfoy_hisseler + ["MANUEL GİRİŞ"])
        else:
            alarm_hisse = al1.selectbox("Hisse", ["MANUEL GİRİŞ"])

        if alarm_hisse == "MANUEL GİRİŞ":
            alarm_hisse = al1.text_input("Hisse Kodu").upper()

        alarm_hedef = al2.number_input("Hedef Fiyat (₺)", min_value=0.0, step=0.01)
        alarm_yon   = al3.selectbox("Koşul", ["Üstüne Çıkınca", "Altına Düşünce"])
        al4.markdown("<div style='margin-top:28px;'></div>", unsafe_allow_html=True)

        if st.form_submit_button("🔔 Alarm Ekle"):
            if alarm_hisse and alarm_hedef > 0:
                st.session_state.alarmlar.append({
                    "Hisse": alarm_hisse,
                    "Hedef": alarm_hedef,
                    "Yon":   alarm_yon,
                    "Aktif": True,
                    "Tarih": datetime.now(pytz.timezone('Europe/Istanbul')).strftime('%d.%m.%Y %H:%M')
                })
                save_json(ALARM_DOSYASI, st.session_state.alarmlar)
                st.success(f"✅ {alarm_hisse} için {tr_format(alarm_hedef)} ₺ alarmı eklendi.")
                st.rerun()

    st.divider()

    if st.session_state.alarmlar:
        st.markdown("#### 📋 Aktif Alarmlar")
        alarm_table = (
            "<table class='kral-table'><thead>"
            "<tr><th>HİSSE</th><th>HEDEF</th><th>KOŞUL</th><th>ANLİK</th><th>DURUM</th><th>EKLENME</th><th>İŞLEM</th></tr>"
            "</thead><tbody>"
        )
        for idx, alarm in enumerate(st.session_state.alarmlar):
            hisse  = alarm.get("Hisse", "-")
            hedef  = alarm.get("Hedef", 0.0)
            yon    = alarm.get("Yon",   "-")
            aktif  = alarm.get("Aktif", True)
            tarih  = alarm.get("Tarih", "-")

            # Anlık fiyatı portföyden bul
            anlik_fiyat = None
            for row in full_data:
                if row["Hisse"] == hisse:
                    anlik_fiyat = row["Güncel"]
                    break

            # Tetiklenme durumu
            tetiklendi = False
            if anlik_fiyat is not None:
                if yon == "Üstüne Çıkınca" and anlik_fiyat >= hedef:
                    tetiklendi = True
                elif yon == "Altına Düşünce" and anlik_fiyat <= hedef:
                    tetiklendi = True

            durum_str = (
                "<span class='alarm-aktif'>🚨 TETİKLENDİ</span>"
                if tetiklendi else
                ("<span style='color:#00e676;'>✅ Aktif</span>" if aktif else "<span style='color:#888;'>⏸ Pasif</span>")
            )
            anlik_str  = tr_format(anlik_fiyat) + " ₺" if anlik_fiyat else "—"
            yon_sembol = "⬆️" if "Üstüne" in yon else "⬇️"

            alarm_table += (
                f"<tr>"
                f"<td><b>{hisse}</b></td>"
                f"<td style='color:{t_sec['accent']};font-weight:bold;'>{tr_format(hedef)} ₺</td>"
                f"<td>{yon_sembol} {yon}</td>"
                f"<td>{anlik_str}</td>"
                f"<td>{durum_str}</td>"
                f"<td style='font-size:11px;'>{tarih}</td>"
                f"<td>—</td>"
                f"</tr>"
            )
        st.markdown(alarm_table + "</tbody></table>", unsafe_allow_html=True)

        # Silme işlemleri (tablo dışında, Streamlit butonları)
        st.markdown("<br>", unsafe_allow_html=True)
        with st.expander("🗑️ Alarm Sil / Duraklat"):
            for idx, alarm in enumerate(st.session_state.alarmlar):
                hisse = alarm.get("Hisse", "-")
                hedef = alarm.get("Hedef", 0.0)
                aktif = alarm.get("Aktif", True)
                bc1, bc2, bc3 = st.columns([3, 1, 1])
                bc1.markdown(f"**{hisse}** → {tr_format(hedef)} ₺ ({alarm.get('Yon', '')})")

                if bc2.button("⏸ Duraklat" if aktif else "▶️ Aktifleştir", key=f"pause_{idx}"):
                    st.session_state.alarmlar[idx]['Aktif'] = not aktif
                    save_json(ALARM_DOSYASI, st.session_state.alarmlar)
                    st.rerun()

                if bc3.button("❌ Sil", key=f"del_alarm_{idx}"):
                    st.session_state.alarmlar.pop(idx)
                    save_json(ALARM_DOSYASI, st.session_state.alarmlar)
                    st.rerun()
    else:
        st.info("Henüz alarm eklenmemiş.")

# ==========================================
# ANALİZ TABU
# ==========================================
with tab_analiz:
    acc = t_sec['accent']
    txt = t_sec['text']
    box = t_sec['box']
    bg  = t_sec['bg']

    # ---------- A) HİSSE FİYAT GRAFİĞİ (Candlestick) ----------
    st.markdown(f"<h4 style='color:{acc};'>🕯️ Hisse Fiyat Grafiği (60 Gün)</h4>", unsafe_allow_html=True)
    portfoy_hisseler_a = sorted(set(x['Hisse'] for x in full_data if x['Piyasa'] == 'Türk Borsası'))
    if portfoy_hisseler_a:
        cs1, cs2 = st.columns([3, 1])
        cs_hisse = cs1.selectbox("Hisse Seç", portfoy_hisseler_a, key="cs_hisse")
        cs_tip   = cs2.selectbox("Grafik Tipi", ["Candlestick", "Çizgi"], key="cs_tip")

        cs_data = fetch_stock_data(cs_hisse)
        if cs_data and not cs_data['hist'].empty:
            hist = cs_data['hist'].copy()
            hist.index = hist.index.tz_localize(None) if hist.index.tz is not None else hist.index

            if cs_tip == "Candlestick":
                cs_fig = go.Figure(data=[go.Candlestick(
                    x=hist.index,
                    open=hist['Open'],
                    high=hist['High'],
                    low=hist['Low'],
                    close=hist['Close'],
                    name=cs_hisse,
                )])
                # Renkleri update_traces ile ayrı ayrı ayarla (versiyon uyumlu)
                cs_fig.update_traces(
                    selector=dict(type='candlestick'),
                    increasing=dict(line=dict(color='#00e676')),
                    decreasing=dict(line=dict(color='#ff1744')),
                )
            else:
                cs_fig = go.Figure(data=[go.Scatter(
                    x=hist.index, y=hist['Close'],
                    mode='lines',
                    line=dict(color=acc, width=2),
                    fill='tozeroy',
                    fillcolor=hex_rgba(acc, 0.09),
                    name=cs_hisse,
                )])

            # MA20 üzerine ekle
            ma20 = hist['Close'].rolling(20).mean()
            cs_fig.add_trace(go.Scatter(
                x=hist.index, y=ma20,
                mode='lines',
                line=dict(color='#ffc107', width=1.2, dash='dot'),
                name='MA20', opacity=0.8,
            ))

            cs_fig.update_layout(
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color=txt, size=11, family=secili_font),
                height=380,
                margin=dict(t=40, b=40, l=40, r=20),
                xaxis=dict(
                    showgrid=True, gridcolor=hex_rgba(acc, 0.08),
                    rangeslider=dict(visible=False),
                    color=txt,
                ),
                yaxis=dict(
                    showgrid=True, gridcolor=hex_rgba(acc, 0.08),
                    color=txt, side='right',
                ),
                legend=dict(
                    orientation='h', x=0, y=1.08,
                    font=dict(size=10, color=txt),
                    bgcolor='rgba(0,0,0,0)',
                ),
                title=dict(
                    text=f"{cs_hisse} — Son 60 Gün",
                    font=dict(size=13, color=acc),
                    x=0.5, xanchor='center',
                ),
            )
            st.plotly_chart(cs_fig, use_container_width=True)
        else:
            st.warning(f"{cs_hisse} için veri alınamadı.")
    else:
        st.info("Türk Borsası'nda hisse bulunamadı.")

    st.divider()

    # ---------- B) PORTFÖY PERFORMANS GRAFİĞİ ----------
    st.markdown(f"<h4 style='color:{acc};'>📉 Portföy Performans Geçmişi</h4>", unsafe_allow_html=True)
    perf_kayitlar = st.session_state.get('performans', [])
    if len(perf_kayitlar) >= 2:
        perf_df = pd.DataFrame(perf_kayitlar)
        perf_df['tarih'] = pd.to_datetime(perf_df['tarih'])
        perf_df = perf_df.sort_values('tarih')

        pf_fig = go.Figure()
        pf_fig.add_trace(go.Scatter(
            x=perf_df['tarih'], y=perf_df['deger'],
            mode='lines+markers',
            name='Portföy Değeri',
            line=dict(color=acc, width=2.5),
            marker=dict(size=5, color=acc),
            fill='tozeroy', fillcolor=hex_rgba(acc, 0.08),
            hovertemplate='<b>%{x|%d.%m.%Y}</b><br>Değer: %{y:,.0f} ₺<extra></extra>',
        ))
        pf_fig.add_trace(go.Bar(
            x=perf_df['tarih'], y=perf_df['kz'],
            name='Günlük K/Z',
            marker_color=[('#00e676' if v >= 0 else '#ff1744') for v in perf_df['kz']],
            opacity=0.6,
            yaxis='y2',
            hovertemplate='<b>%{x|%d.%m.%Y}</b><br>K/Z: %{y:,.0f} ₺<extra></extra>',
        ))
        pf_fig.update_layout(
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)',
            font=dict(color=txt, size=11, family=secili_font),
            height=360,
            margin=dict(t=50, b=40, l=40, r=60),
            xaxis=dict(showgrid=True, gridcolor=hex_rgba(acc, 0.08), color=txt),
            yaxis=dict(showgrid=True, gridcolor=hex_rgba(acc, 0.08), color=txt,
                       title=dict(text='Değer (₺)', font=dict(size=11))),
            yaxis2=dict(overlaying='y', side='right', showgrid=False, color=txt,
                        title=dict(text='K/Z (₺)', font=dict(size=11))),
            legend=dict(orientation='h', x=0, y=1.08,
                        font=dict(size=10, color=txt), bgcolor='rgba(0,0,0,0)'),
            title=dict(text="Günlük Portföy Değeri & K/Z",
                       font=dict(size=13, color=acc), x=0.5, xanchor='center'),
            barmode='overlay',
        )
        st.plotly_chart(pf_fig, use_container_width=True)

        # Özet istatistikler
        _ilk  = perf_df['deger'].iloc[0]
        _son  = perf_df['deger'].iloc[-1]
        _max  = perf_df['deger'].max()
        _min  = perf_df['deger'].min()
        _perf = ((_son - _ilk) / _ilk * 100) if _ilk > 0 else 0
        _pc   = "#00e676" if _perf >= 0 else "#ff1744"
        ps1, ps2, ps3, ps4 = st.columns(4)
        ps1.metric("Dönem Başı",        f"{tr_format(_ilk)} ₺")
        ps2.metric("Bugün",             f"{tr_format(_son)} ₺")
        ps3.metric("Dönem Getirisi",    f"%{_perf:+.2f}")
        ps4.metric("En Yüksek Değer",   f"{tr_format(_max)} ₺")
    else:
        st.info("Performans geçmişi için en az 2 günlük veri gerekli. Uygulama her açıldığında o günün değerini kaydeder.")

    st.divider()

    # ---------- C) KORELASYON MATRİSİ ----------
    st.markdown(f"<h4 style='color:{acc};'>🔗 Hisse Korelasyon Matrisi</h4>", unsafe_allow_html=True)
    portfoy_bist = [x['Hisse'] for x in full_data if x['Piyasa'] == 'Türk Borsası']
    if len(portfoy_bist) >= 2:
        kor_df = hesapla_korelasyon(tuple(portfoy_bist))
        if kor_df is not None and not kor_df.empty:
            n = len(kor_df)
            kor_z = kor_df.values
            labels = kor_df.columns.tolist()

            # Renk skalası: kırmızı(-1) → beyaz(0) → yeşil(+1)
            kor_fig = go.Figure(data=go.Heatmap(
                z=kor_z,
                x=labels,
                y=labels,
                colorscale=[
                    [0.0,  '#ff1744'],
                    [0.5,  box],
                    [1.0,  '#00e676'],
                ],
                zmin=-1, zmax=1,
                text=[[f"{kor_z[i][j]:.2f}" for j in range(n)] for i in range(n)],
                texttemplate="%{text}",
                textfont=dict(size=11, color=txt),
                hovertemplate='<b>%{y} × %{x}</b><br>Korelasyon: %{z:.3f}<extra></extra>',
                showscale=True,
                colorbar=dict(
                    thickness=12, len=0.8,
                    tickfont=dict(size=10, color=txt),
                    title=dict(text='r', font=dict(color=txt)),
                ),
            ))
            kor_fig.update_layout(
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color=txt, size=11, family=secili_font),
                height=max(300, n * 44 + 80),
                margin=dict(t=50, b=60, l=60, r=80),
                xaxis=dict(color=txt, tickangle=-35),
                yaxis=dict(color=txt),
                title=dict(
                    text="Son 30 Gün Kapanış Korelasyonu",
                    font=dict(size=13, color=acc),
                    x=0.5, xanchor='center',
                ),
            )
            st.plotly_chart(kor_fig, use_container_width=True)
            st.markdown(
                f"<small style='color:{acc}88;'>+1,00 = Tam pozitif korelasyon &nbsp;|&nbsp; "
                f"0,00 = İlişkisiz &nbsp;|&nbsp; -1,00 = Tam negatif korelasyon</small>",
                unsafe_allow_html=True
            )
        else:
            st.warning("Korelasyon hesaplanamadı (yeterli veri yok).")
    else:
        st.info("Korelasyon matrisi için portföyde en az 2 Türk Borsası hissesi gerekli.")

# ==========================================
# HABERLER TABU
# ==========================================
with tab_haberler:
    acc = t_sec['accent']; txt = t_sec['text']; box = t_sec['box']
    st.markdown(f"<h4 style='color:{acc};'>📰 Portföy Haber Akışı</h4>", unsafe_allow_html=True)
    st.markdown(
        f"<small style='color:{acc}88;'>Portföydeki hisselere ait son haberler — yfinance üzerinden</small>",
        unsafe_allow_html=True
    )

    portfoy_hisseleri_haber = sorted(set(
        x['Hisse'] for x in full_data if x['Piyasa'] == 'Türk Borsası'
    ))

    if portfoy_hisseleri_haber:
        with st.spinner("Haberler yükleniyor..."):
            haberler = fetch_haberler(tuple(portfoy_hisseleri_haber))

        if haberler:
            # Hisse filtresi
            haber_filtre = st.multiselect(
                "Hisse Filtrele",
                ["Tümü"] + [h.replace(".IS", "") for h in portfoy_hisseleri_haber],
                default=["Tümü"],
                key="haber_filtre"
            )
            gosterilecek = haberler
            if "Tümü" not in haber_filtre and haber_filtre:
                gosterilecek = [h for h in haberler if h['hisse'] in haber_filtre]

            for haber in gosterilecek[:25]:
                zaman_str = ""
                if haber['zaman']:
                    try:
                        dt = datetime.fromtimestamp(haber['zaman'], tz=pytz.timezone('Europe/Istanbul'))
                        zaman_str = dt.strftime('%d.%m.%Y %H:%M')
                    except Exception:
                        pass

                st.markdown(
                    f"<div style='background:{box};border:1px solid {acc}22;border-radius:8px;"
                    f"padding:12px 16px;margin-bottom:8px;'>"
                    f"<div style='display:flex;justify-content:space-between;align-items:flex-start;gap:10px;'>"
                    f"  <div>"
                    f"    <span style='background:{acc}22;color:{acc};font-size:10px;font-weight:700;"
                    f"          padding:2px 7px;border-radius:4px;margin-right:8px;'>{haber['hisse']}</span>"
                    f"    <a href='{haber['url']}' target='_blank' style='color:{txt};text-decoration:none;"
                    f"       font-size:13px;font-weight:600;'>{haber['baslik']}</a>"
                    f"  </div>"
                    f"</div>"
                    f"<div style='margin-top:6px;display:flex;gap:12px;'>"
                    f"  <span style='font-size:10px;opacity:0.45;'>{haber['kaynak']}</span>"
                    f"  <span style='font-size:10px;opacity:0.45;'>{zaman_str}</span>"
                    f"</div>"
                    f"</div>",
                    unsafe_allow_html=True
                )
        else:
            st.info("Haber bulunamadı. yfinance bazı semboller için haber döndürmeyebilir.")
    else:
        st.info("Haber akışı için portföyde Türk Borsası hissesi bulunmalıdır.")


# ==========================================
# TEMEL VERİLER TABU
# ==========================================
with tab_temel:
    acc = t_sec['accent']; txt = t_sec['text']; box = t_sec['box']
    st.markdown(f"<h4 style='color:{acc};'>🏦 Temel Analiz Verileri</h4>", unsafe_allow_html=True)
    st.markdown(
        f"<small style='color:{acc}88;'>F/K, PD/DD, Borç/Özsermaye — yfinance üzerinden · Veriler yaklaşık değerdir</small>",
        unsafe_allow_html=True
    )

    portfoy_bist_temel = [x for x in full_data if x['Piyasa'] == 'Türk Borsası']

    if portfoy_bist_temel:
        with st.spinner("Temel veriler çekiliyor..."):
            temel_rows = []
            for item in portfoy_bist_temel:
                tv = fetch_temel_veri(item['Hisse'])
                temel_rows.append({**item, 'temel': tv or {}})

        # Tablo
        tbl = (
            "<table class='kral-table'><thead><tr>"
            "<th>HİSSE</th><th>F/K</th><th>PD/DD</th>"
            "<th>BORÇ/ÖZSERM.</th><th>PİYASA DEĞERİ</th>"
            "<th>TEMETTÜ VERİMİ</th><th>GÜNCEL FİYAT</th>"
            "</tr></thead><tbody>"
        )
        for r in temel_rows:
            tv = r['temel']

            def fmt_val(v, suffix=''):
                if v is None: return "<span style='opacity:0.35;'>—</span>"
                return f"{v}{suffix}"

            def fmt_mkt(v):
                if v is None: return "<span style='opacity:0.35;'>—</span>"
                if v >= 1e9:  return f"{v/1e9:.1f} Mr ₺"
                if v >= 1e6:  return f"{v/1e6:.0f} Mn ₺"
                return f"{v:,.0f} ₺"

            # Borç/Özsermaye rengi
            bo = tv.get('Borc_Ozserm')
            bo_color = '#00e676' if bo is not None and bo < 0.5 else ('#ffc107' if bo is not None and bo < 1.5 else '#ff1744')
            bo_str = f"<span style='color:{bo_color};font-weight:600;'>{bo}</span>" if bo is not None else "<span style='opacity:0.35;'>—</span>"

            # F/K rengi
            fk = tv.get('FK')
            fk_color = '#00e676' if fk is not None and fk < 15 else ('#ffc107' if fk is not None and fk < 30 else '#ff1744')
            fk_str = f"<span style='color:{fk_color};font-weight:600;'>{fk}</span>" if fk is not None else "<span style='opacity:0.35;'>—</span>"

            tbl += (
                f"<tr>"
                f"<td><b>{r['Hisse']}</b></td>"
                f"<td>{fk_str}</td>"
                f"<td>{fmt_val(tv.get('PD_DD'))}</td>"
                f"<td>{bo_str}</td>"
                f"<td style='font-size:11px;'>{fmt_mkt(tv.get('Piyasa_Degeri'))}</td>"
                f"<td>{'%'+str(tv.get('Temettu_Verimi')) if tv.get('Temettu_Verimi') else '<span style=\"opacity:0.35;\">—</span>'}</td>"
                f"<td>{tr_format4(r['Güncel'])} ₺</td>"
                f"</tr>"
            )
        tbl += "</tbody></table>"
        st.markdown(tbl, unsafe_allow_html=True)

        # Açıklama kartı
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(
            f"<div style='background:{box};border:1px solid {acc}22;border-radius:8px;padding:12px 16px;"
            f"display:grid;grid-template-columns:1fr 1fr;gap:8px;font-size:11px;'>"
            f"<div><b style='color:{acc};'>F/K (Fiyat/Kazanç)</b><br>"
            f"<span style='opacity:0.6;'>&lt;15 ucuz · 15–30 makul · &gt;30 pahalı</span></div>"
            f"<div><b style='color:{acc};'>PD/DD (Piyasa/Defter)</b><br>"
            f"<span style='opacity:0.6;'>&lt;1 defter altı · 1–3 makul · &gt;3 prim</span></div>"
            f"<div><b style='color:#00e676;'>Borç/Özserm. &lt;0.5</b> düşük · "
            f"<b style='color:#ffc107;'>0.5–1.5</b> orta · "
            f"<b style='color:#ff1744;'>&gt;1.5</b> yüksek</div>"
            f"<div style='opacity:0.5;'>Veriler yfinance üzerinden çekilmekte olup kesin değildir.</div>"
            f"</div>",
            unsafe_allow_html=True
        )

        # BIST100 Karşılaştırma grafiği
        st.divider()
        st.markdown(f"<h4 style='color:{acc};'>📊 Portföy vs BIST100 (60 Gün Normalize)</h4>", unsafe_allow_html=True)
        bist_hisseler = tuple(x['Hisse'] for x in portfoy_bist_temel)
        with st.spinner("Karşılaştırma grafiği hazırlanıyor..."):
            bist_norm, portfoy_norm = fetch_bist100_karsilastirma(bist_hisseler)

        if bist_norm is not None:
            kars_fig = go.Figure()
            kars_fig.add_trace(go.Scatter(
                x=bist_norm.index, y=bist_norm.values,
                mode='lines', name='BIST 100',
                line=dict(color='#ffc107', width=2, dash='dot'),
                hovertemplate='<b>BIST100</b><br>%{x|%d.%m.%Y}<br>%{y:.1f}<extra></extra>',
            ))
            if portfoy_norm is not None:
                kars_fig.add_trace(go.Scatter(
                    x=portfoy_norm.index, y=portfoy_norm.values,
                    mode='lines', name='Portföy Ort.',
                    line=dict(color=acc, width=2.5),
                    fill='tonexty', fillcolor=hex_rgba(acc, 0.06),
                    hovertemplate='<b>Portföy</b><br>%{x|%d.%m.%Y}<br>%{y:.1f}<extra></extra>',
                ))
            kars_fig.add_hline(y=100, line_dash='dash', line_color=txt, opacity=0.2)
            kars_fig.update_layout(
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color=txt, size=11, family=secili_font),
                height=360,
                margin=dict(t=40, b=40, l=50, r=20),
                xaxis=dict(showgrid=True, gridcolor=hex_rgba(acc, 0.08), color=txt),
                yaxis=dict(showgrid=True, gridcolor=hex_rgba(acc, 0.08), color=txt,
                           title=dict(text='Normalize (Başlangıç=100)', font=dict(size=10))),
                legend=dict(orientation='h', x=0, y=1.08,
                            font=dict(size=10, color=txt), bgcolor='rgba(0,0,0,0)'),
                title=dict(text="Portföy Ortalama vs BIST100",
                           font=dict(size=13, color=acc), x=0.5, xanchor='center'),
            )
            st.plotly_chart(kars_fig, use_container_width=True)

            # Beta hesapla
            if portfoy_norm is not None and len(bist_norm) > 10:
                try:
                    bist_ret = bist_norm.pct_change().dropna()
                    port_ret = portfoy_norm.pct_change().dropna()
                    min_l = min(len(bist_ret), len(port_ret))
                    cov    = np.cov(port_ret.values[-min_l:], bist_ret.values[-min_l:])
                    beta   = round(cov[0, 1] / cov[1, 1], 2) if cov[1, 1] != 0 else None
                    if beta is not None:
                        beta_color = '#00e676' if beta < 1 else ('#ffc107' if beta < 1.5 else '#ff1744')
                        st.markdown(
                            f"<div style='text-align:center;margin-top:4px;'>"
                            f"<span style='font-size:12px;opacity:0.6;'>Portföy Betası: </span>"
                            f"<span style='color:{beta_color};font-weight:700;font-size:16px;'>β = {beta}</span>"
                            f"<span style='font-size:11px;opacity:0.45;'> "
                            f"({'düşük risk' if beta < 1 else ('orta risk' if beta < 1.5 else 'yüksek risk')})</span>"
                            f"</div>",
                            unsafe_allow_html=True
                        )
                except Exception:
                    pass
        else:
            st.info("Karşılaştırma verisi alınamadı.")
    else:
        st.info("Türk Borsası'nda hisse bulunamadı.")


# ==========================================
# NOTLAR TABU
# ==========================================
with tab_notlar:
    acc = t_sec['accent']; txt = t_sec['text']; box = t_sec['box']
    st.markdown(f"<h4 style='color:{acc};'>📝 Hisse Notları & Hedef Fiyatlar</h4>", unsafe_allow_html=True)

    portfoy_hisseleri_not = sorted(set(x['Hisse'] for x in full_data))

    if portfoy_hisseleri_not:
        not_hisse = st.selectbox("Hisse Seç", portfoy_hisseleri_not, key="not_hisse_sec")
        mevcut    = st.session_state.notlar.get(not_hisse, {})
        guncel_fiyat = next((x['Güncel'] for x in full_data if x['Hisse'] == not_hisse), 0.0)

        with st.form(f"not_form_{not_hisse}", clear_on_submit=False):
            nf1, nf2 = st.columns(2)
            hedef_fiyat = nf1.number_input(
                "Hedef Fiyat (₺)",
                value=float(mevcut.get('hedef', 0.0)),
                format="%.4f", min_value=0.0,
                key=f"hedef_{not_hisse}"
            )
            stop_loss = nf2.number_input(
                "Stop-Loss (₺)",
                value=float(mevcut.get('stop', 0.0)),
                format="%.4f", min_value=0.0,
                key=f"stop_{not_hisse}"
            )
            alim_sebebi = st.text_area(
                "Alım Gerekçesi",
                value=mevcut.get('sebep', ''),
                placeholder="Bu hisseyi neden aldım? Hangi beklentilerle?",
                height=100, key=f"sebep_{not_hisse}"
            )
            genel_not = st.text_area(
                "Genel Not",
                value=mevcut.get('not', ''),
                placeholder="Takip ettiğim gelişmeler, riskler, çıkış stratejisi...",
                height=80, key=f"gnot_{not_hisse}"
            )
            if st.form_submit_button("💾 Notu Kaydet", use_container_width=True):
                st.session_state.notlar[not_hisse] = {
                    'Hisse': not_hisse,
                    'hedef': hedef_fiyat,
                    'stop':  stop_loss,
                    'sebep': alim_sebebi,
                    'not':   genel_not,
                    'guncelleme': datetime.now(pytz.timezone('Europe/Istanbul')).strftime('%d.%m.%Y %H:%M'),
                }
                save_json(NOTLAR_DOSYASI, list(st.session_state.notlar.values()))
                st.success(f"✅ {not_hisse} notu kaydedildi.")

        # Hedef & stop bilgi kartı
        if hedef_fiyat > 0 or stop_loss > 0:
            pot_kar  = ((hedef_fiyat - guncel_fiyat) / guncel_fiyat * 100) if guncel_fiyat > 0 and hedef_fiyat > 0 else None
            pot_risk = ((guncel_fiyat - stop_loss)   / guncel_fiyat * 100) if guncel_fiyat > 0 and stop_loss > 0 else None
            rr_oran  = round(pot_kar / pot_risk, 2) if pot_kar and pot_risk and pot_risk > 0 else None
            st.markdown(
                f"<div style='background:{box};border:1px solid {acc}33;border-radius:10px;"
                f"padding:14px 16px;margin-top:8px;"
                f"display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:10px;'>"
                f"<div><div style='font-size:10px;opacity:0.5;'>GÜNCEL</div>"
                f"    <div style='color:{acc};font-weight:700;'>{tr_format(guncel_fiyat)} ₺</div></div>"
                f"<div><div style='font-size:10px;opacity:0.5;'>HEDEF</div>"
                f"    <div style='color:#00e676;font-weight:700;'>{tr_format(hedef_fiyat)} ₺</div>"
                f"    {'<div style=\"font-size:10px;color:#00e676;\">+'+str(round(pot_kar,1))+'%</div>' if pot_kar else ''}"
                f"</div>"
                f"<div><div style='font-size:10px;opacity:0.5;'>STOP-LOSS</div>"
                f"    <div style='color:#ff1744;font-weight:700;'>{tr_format(stop_loss)} ₺</div>"
                f"    {'<div style=\"font-size:10px;color:#ff1744;\">-'+str(round(pot_risk,1))+'%</div>' if pot_risk else ''}"
                f"</div>"
                f"<div><div style='font-size:10px;opacity:0.5;'>R/R ORANI</div>"
                f"    <div style='color:{'#00e676' if rr_oran and rr_oran >= 2 else '#ffc107'};font-weight:700;'>"
                f"    {'1 : '+str(rr_oran) if rr_oran else '—'}</div></div>"
                f"</div>",
                unsafe_allow_html=True
            )

        # Tüm notlar özeti
        st.divider()
        st.markdown(f"#### 📋 Tüm Hisse Notları", unsafe_allow_html=True)
        if st.session_state.notlar:
            for h, n in sorted(st.session_state.notlar.items()):
                if n.get('sebep') or n.get('hedef') or n.get('not'):
                    with st.expander(f"📌 {h}  |  Hedef: {tr_format(n.get('hedef',0))} ₺  |  Stop: {tr_format(n.get('stop',0))} ₺"):
                        if n.get('sebep'):
                            st.markdown(f"**Alım Gerekçesi:** {n['sebep']}")
                        if n.get('not'):
                            st.markdown(f"**Not:** {n['not']}")
                        st.caption(f"Son güncelleme: {n.get('guncelleme', '—')}")
        else:
            st.info("Henüz not girilmemiş.")
    else:
        st.info("Portföyde hisse bulunmamaktadır.")


# ==========================================
# DIŞA AKTAR TABU
# ==========================================
with tab_export:
    acc = t_sec['accent']; txt = t_sec['text']; box = t_sec['box']
    st.markdown(f"<h4 style='color:{acc};'>📤 Portföy Dışa Aktarma</h4>", unsafe_allow_html=True)

    if full_data:
        exp_df = pd.DataFrame([{
            'Hisse':        r['Hisse'],
            'Piyasa':       r['Piyasa'],
            'Adet':         r['Adet'],
            'Maliyet (₺)':  r['Maliyet'],
            'Güncel (₺)':   r['Güncel'],
            'K/Z (₺)':      r['K/Z'],
            'Toplam Değer': r['Değer'],
            'Günlük Değ.':  r['DailyDiff'],
            'Net Temettü':  r['NetTemettu'],
            'Sinyal':       r['Sinyal'],
            'RSI':          r['RSI'],
        } for r in full_data])

        # Özet istatistikler
        e1, e2, e3, e4 = st.columns(4)
        e1.metric("Toplam Değer",    f"{tr_format(exp_df['Toplam Değer'].sum())} ₺")
        e2.metric("Toplam K/Z",      f"{tr_format(exp_df['K/Z (₺)'].sum())} ₺")
        e3.metric("Yıllık Temettü",  f"{tr_format(exp_df['Net Temettü'].sum())} ₺")
        e4.metric("Pozisyon Sayısı", len(exp_df))

        st.divider()
        st.markdown("#### CSV İndir")

        # CSV
        csv_data = exp_df.to_csv(index=False, encoding='utf-8-sig', sep=';', decimal=',')
        st.download_button(
            label="⬇️ CSV Olarak İndir",
            data=csv_data.encode('utf-8-sig'),
            file_name=f"portfoy_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime='text/csv',
            use_container_width=True,
        )

        st.markdown("#### Excel İndir")
        # Excel (in-memory, openpyxl)
        try:
            import io
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Portföy sayfası
                exp_df.to_excel(writer, index=False, sheet_name='Portföy')
                ws = writer.sheets['Portföy']

                # Başlık stili
                hdr_fill = PatternFill(fill_type='solid', fgColor='1B2838')
                hdr_font = Font(color='00D4FF', bold=True, size=11)
                for cell in ws[1]:
                    cell.fill      = hdr_fill
                    cell.font      = hdr_font
                    cell.alignment = Alignment(horizontal='center')

                # Kolon genişlikleri
                for col in ws.columns:
                    max_w = max(len(str(c.value or '')) for c in col) + 4
                    ws.column_dimensions[col[0].column_letter].width = min(max_w, 22)

                # K/Z renklendirme
                kz_col = [c.column_letter for c in ws[1] if c.value == 'K/Z (₺)']
                if kz_col:
                    for row in ws.iter_rows(min_row=2, min_col=ws[1][exp_df.columns.tolist().index('K/Z (₺)')].column, max_col=ws[1][exp_df.columns.tolist().index('K/Z (₺)')].column):
                        for cell in row:
                            try:
                                v = float(cell.value or 0)
                                cell.font = Font(color='00E676' if v >= 0 else 'FF1744', bold=True)
                            except Exception:
                                pass

                # Performans sayfası
                if st.session_state.performans:
                    perf_df_exp = pd.DataFrame(st.session_state.performans)
                    perf_df_exp.to_excel(writer, index=False, sheet_name='Performans Geçmişi')

            excel_bytes = output.getvalue()
            st.download_button(
                label="⬇️ Excel Olarak İndir (.xlsx)",
                data=excel_bytes,
                file_name=f"portfoy_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True,
            )
        except ImportError:
            st.warning("Excel export için `openpyxl` paketi gerekli: `pip install openpyxl`")
        except Exception as ex:
            st.error(f"Excel oluşturma hatası: {ex}")

        # Önizleme
        st.divider()
        st.markdown("#### Önizleme")
        st.dataframe(exp_df, use_container_width=True, hide_index=True)

    else:
        st.info("Dışa aktarmak için portföyde varlık bulunmalıdır.")

# ==========================================
# FOOTER
# ==========================================
st.markdown("---")
st.caption(
    f"🕒 Son Güncelleme: {datetime.now(pytz.timezone('Europe/Istanbul')).strftime('%H:%M:%S')}  |  "
    f"Sinyal: RSI + MA20 + MACD + Bollinger Bands  |  * Son 1 yılda temettü dağıtımı yapılmamış hisseler"
)
